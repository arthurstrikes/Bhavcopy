"""
NAV Calculator - IMP Portfolios.  Place at pages/nav_calculator.py

Price source: official NSE bhavcopy (nse_bhavcopy.py at repo root).
Engine:       imp_engine.py at repo root.
Factsheet:    pdf_report.py at repo root.

Design shell: INTERNAL suite - generic blue #1E4FD8, Plus Jakarta Sans + Inter.
MOFSL brand colours are reserved for the external/IFA shell and are deliberately
NOT used here. Green/red/amber are universal signal colours and are never swapped.
To brand-swap later, change PALETTE only: #1E4FD8 -> MOFSL navy, tint8 -> primary
at 8%, tint30 -> primary at 30%.

Streamlit 1.39.0 API only - use_container_width, never width=.
"""

import hashlib
import io
from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
import yfinance as yf

import imp_engine as engine
import mtf_leverage as mtf
import nse_bhavcopy

try:
    import pdf_report
    PDF_OK = True
except Exception:
    PDF_OK = False

# -- design tokens ------------------------------------------------------------
# Single source of truth. The theme in .streamlit/config.toml must mirror
# primary / bg / surface / ink or the Streamlit chrome will not match the cards.

PALETTE = {
    "primary": "#1E4FD8",
    "primary_dk": "#173CA6",
    "tint8": "#EBF0FF",
    "tint30": "#B8CAFF",
    "pos": "#0F8047",
    "neg": "#C0392B",
    "amber": "#B87800",
    "ink": "#12161F",
    "muted": "#5B6472",
    "rule": "#E3E6EB",
    "surface": "#F7F8FA",
    "bg": "#FFFFFF",
}
P = PALETTE

st.set_page_config(page_title="NAV Calculator - IMP Portfolios",
                   page_icon="chart_with_upwards_trend", layout="wide")

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@500;600;700&family=Inter:wght@400;500;600&display=swap');

:root {{
  --primary:{P['primary']}; --primary-dk:{P['primary_dk']};
  --tint8:{P['tint8']}; --tint30:{P['tint30']};
  --pos:{P['pos']}; --neg:{P['neg']}; --amber:{P['amber']};
  --ink:{P['ink']}; --muted:{P['muted']}; --rule:{P['rule']};
  --surface:{P['surface']}; --bg:{P['bg']};
  --s1:4px; --s2:8px; --s3:16px; --s4:24px; --s5:32px;
}}

html, body, [class*="css"], .stMarkdown, p, li, label {{
  font-family:'Inter',system-ui,sans-serif;
}}
h1,h2,h3,h4,h5 {{ font-family:'Plus Jakarta Sans','Inter',sans-serif; }}

/* Tabular figures wherever numbers appear - decimals must line up. */
[data-testid="stDataFrame"], .kpi-val, .focal-val, [data-testid="stMetricValue"] {{
  font-variant-numeric:tabular-nums; font-feature-settings:'tnum' 1;
}}

.block-container {{ padding-top:2.2rem; padding-bottom:3rem; max-width:1400px; }}

/* report header */
.rh-title {{ font-family:'Plus Jakarta Sans','Inter',-apple-system,
                        BlinkMacSystemFont,'Segoe UI',sans-serif;
            font-size:1.55rem; font-weight:700; color:var(--ink);
            letter-spacing:-.02em; line-height:1.3; padding-top:2px; margin:0; }}
.rh-strip {{ display:flex; flex-wrap:wrap; gap:0 var(--s4); margin:var(--s2) 0 var(--s4);
             padding-bottom:var(--s3); border-bottom:1px solid var(--rule); }}
.rh-item {{ display:flex; flex-direction:column; gap:2px; }}
.rh-k {{ font-size:.63rem; text-transform:uppercase; letter-spacing:.07em;
         color:var(--muted); font-weight:500; }}
.rh-v {{ font-size:.82rem; color:var(--ink); font-weight:600;
         font-family:'Plus Jakarta Sans',sans-serif; }}

/* focal metric + kpi cards */
.focal-val {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:3rem;
              font-weight:700; line-height:1; letter-spacing:-.03em; }}
.focal-lbl {{ font-size:.72rem; text-transform:uppercase; letter-spacing:.07em;
              color:var(--muted); font-weight:600; margin-top:6px; }}
.focal-sub {{ font-size:.8rem; color:var(--muted); }}

.kpi {{ background:var(--bg); border:1px solid var(--rule); border-radius:6px;
        padding:12px 14px; }}
.kpi-val {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:1.15rem;
            font-weight:700; line-height:1.2; letter-spacing:-.015em; }}
.kpi-lbl {{ font-size:.63rem; text-transform:uppercase; letter-spacing:.06em;
            color:var(--muted); margin-top:3px; font-weight:500; }}

.gain {{ color:var(--pos); }} .loss {{ color:var(--neg); }} .neutral {{ color:var(--ink); }}

/* status chip - alerts demoted from a full-width banner */
.chip {{ display:inline-flex; align-items:center; gap:7px; padding:5px 11px;
         border-radius:20px; font-size:.75rem; font-weight:500; }}
.chip-ok {{ background:rgba(15,128,71,.09); color:var(--pos); }}
.chip-warn {{ background:rgba(184,120,0,.10); color:var(--amber); }}
.dot {{ width:6px; height:6px; border-radius:50%; background:currentColor; }}

.sec {{ font-family:'Plus Jakarta Sans',sans-serif; font-size:.72rem;
        text-transform:uppercase; letter-spacing:.08em; color:var(--muted);
        font-weight:600; margin:var(--s4) 0 var(--s2); }}

.note {{ background:var(--tint8); border-left:3px solid var(--primary);
         padding:12px 14px; border-radius:0 5px 5px 0; font-size:.82rem;
         color:var(--ink); margin-bottom:var(--s2); }}
.warn {{ background:rgba(184,120,0,.08); border-left:3px solid var(--amber);
         padding:12px 14px; border-radius:0 5px 5px 0; font-size:.82rem;
         color:var(--ink); margin-bottom:var(--s2); }}
.err {{ background:rgba(192,57,43,.07); border-left:3px solid var(--neg);
        padding:12px 14px; border-radius:0 5px 5px 0; font-size:.85rem;
        margin-bottom:var(--s2); }}

.foot {{ font-size:.68rem; color:var(--muted); line-height:1.55;
         border-top:1px solid var(--rule); padding-top:var(--s3);
         margin-top:var(--s5); }}

/* streamlit chrome */
section[data-testid="stSidebar"] {{ background:var(--surface);
                                    border-right:1px solid var(--rule); }}
section[data-testid="stSidebar"] h3 {{ font-size:.72rem !important;
    text-transform:uppercase; letter-spacing:.08em; color:var(--muted) !important;
    font-weight:600 !important; margin-bottom:var(--s2) !important; }}
.stTabs [data-baseweb="tab-list"] {{ gap:var(--s4); border-bottom:1px solid var(--rule); }}
.stTabs [data-baseweb="tab"] {{ height:38px; padding:0; font-size:.82rem;
                                font-weight:500; color:var(--muted); }}
.stTabs [aria-selected="true"] {{ color:var(--primary) !important; font-weight:600; }}
div[data-testid="stDataFrame"] {{ border:1px solid var(--rule); border-radius:6px; }}
hr {{ border-color:var(--rule); margin:var(--s4) 0; }}
</style>
""", unsafe_allow_html=True)

CAPITAL_OPTIONS = {
    "Rs 1,00,000": 100000, "Rs 1,50,000": 150000, "Rs 2,50,000": 250000,
    "Rs 5,00,000": 500000, "Rs 10,00,000": 1000000, "Rs 25,00,000": 2500000,
    "Rs 50,00,000": 5000000,
}
DEFAULT_CAPITAL = "Rs 2,50,000"

# Indices resolved natively via nse_bhavcopy.INDEX_ALIASES (verified against a
# live NSE index bhavcopy file before being added - see resolve_index_name).
# Anything NOT in this list falls through to BENCH_MAP / yfinance below.
NSE_NATIVE_BENCHMARKS = [
    "Nifty 50", "Nifty 100", "Nifty 200", "Nifty 500", "Nifty Total Market",
    "Nifty Midcap 100", "Nifty Midcap 150", "Nifty Smallcap 100",
    "Nifty Smallcap 250", "Nifty LargeMidcap 250", "Nifty MidSmallcap 400",
    "Nifty200 Momentum 30",
]

# BSE indices are not published by NSE and stay on yfinance permanently (see
# nse_bhavcopy.py). ^BSE500 is UNVERIFIED from this build environment - Yahoo
# Finance access was blocked in the sandbox used to build this, so the ticker
# could not be tested end-to-end. Confirm it resolves on first live use; if it
# returns no data, check the correct BSE 500 ticker on Yahoo Finance directly
# and update BENCH_MAP.
BENCH_MAP = {
    "Sensex": "^BSESN",
    "BSE 500": "^BSE500",  # UNVERIFIED - confirm on first live use
}


class _PastedLog(io.BytesIO):
    """Wraps a text-area paste as a file-like object so engine.load_log() -
    which already sniffs tab-separated text via _read_csv_any_delimiter -
    handles it identically to an uploaded file. Only `.name` needs to exist."""
    name = "pasted_log.csv"


def fmt_inr(v):
    if abs(v) >= 1e7:
        return f"Rs {v/1e7:.2f} Cr"
    if abs(v) >= 1e5:
        return f"Rs {v/1e5:.2f} L"
    return f"Rs {v:,.0f}"


def fmt_pct(v):
    return f"{'+' if v >= 0 else ''}{v:.2f}%"


def tone(v):
    return "gain" if v >= 0 else "loss"


# -- price fetch --------------------------------------------------------------

@st.cache_data(show_spinner=False, ttl=7 * 24 * 3600)
def fetch_closes(start: date, end: date):
    """
    {date: {SYMBOL: close}} for every NSE session in range.

    Every calendar day is probed. Weekends are NOT pre-filtered: NSE runs
    special weekend sessions (Sunday 01-Feb-2026 traded). A 404 from the archive
    is the authoritative non-trading-day signal.

    A determinate progress bar replaces the indefinite spinner - on a first run
    this loop makes ~200 requests and silence reads as a hang.
    """
    days = engine.calendar_days(start, end)
    out, errors = {}, {}

    def one(d):
        try:
            m, traded = nse_bhavcopy.fetch_day(d)
            return d, (m if traded else None), None
        except nse_bhavcopy.BhavcopyFetchError as e:
            return d, None, str(e)[:160]

    bar = st.progress(0.0, text=f"Fetching NSE sessions 0 / {len(days)}")
    done = 0
    with ThreadPoolExecutor(max_workers=nse_bhavcopy.MAX_WORKERS) as ex:
        for d, m, err in ex.map(one, days):
            done += 1
            if done % 5 == 0 or done == len(days):
                bar.progress(done / len(days),
                             text=f"Fetching NSE sessions {done} / {len(days)}")
            if err:
                errors[d] = err
            elif m:
                out[d] = m
    bar.empty()
    return out, errors


def benchmark_series(closes, name, start, end):
    """
    Benchmark read from the bhavcopy files ALREADY fetched.

    nse_bhavcopy.fetch_day returns equities and indices merged, so every index
    close is present from the first pass. Re-fetching cost a second full sweep
    of the archive and, when that sweep hiccuped, silently fell through to
    yfinance - which rate-limits Streamlit Cloud's shared IPs. Names in
    BENCH_MAP (Sensex, BSE 500) need the network, being BSE indices NSE does
    not publish - matched by dict membership, not a hardcoded name, so a new
    BSE index only needs adding to BENCH_MAP, not a second branch here.
    """
    if name == "None":
        return {}
    if name in BENCH_MAP:
        try:
            df = yf.download(BENCH_MAP[name], start=start.strftime("%Y-%m-%d"),
                             end=(end + timedelta(days=2)).strftime("%Y-%m-%d"),
                             interval="1d", progress=False, auto_adjust=False,
                             actions=False)
            if df.empty:
                return {}
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            cc = next((c for c in df.columns if "close" in c.lower()), None)
            if cc is None:
                return {}
            out = {}
            for idx, row in df.iterrows():
                try:
                    v = float(row[cc])
                    if not np.isnan(v):
                        out[idx.date()] = v
                except Exception:
                    pass
            return out
        except Exception:
            return {}

    key = nse_bhavcopy.resolve_index_name(name) or name.upper()
    return {d: m[key] for d, m in closes.items() if key in m and m[key] > 0}


@st.cache_data(show_spinner=False, ttl=7 * 24 * 3600)
def fetch_ca(days):
    """NSE corporate-action calendar, same archive host as the bhavcopy."""
    import time as _t
    sess = requests.Session()
    sess.headers.update(nse_bhavcopy.HEADERS)

    def get(url):
        r = None
        for a in range(5):
            r = sess.get(url, timeout=40)
            if r.status_code not in (403, 503, 429):
                return r
            _t.sleep(1.5 * (a + 1))
        return r

    return engine.fetch_ca_calendar(days, get).to_dict("records")


def ca_sample_days(sessions, every=3):
    """
    NSE lists an action in the PR file for at least 7 calendar days before its
    ex-date - measured across every action in the reference period, the shortest
    lead was 7 days, roughly 5 sessions. Sampling every 3rd session therefore
    cannot miss one, and cuts the download from ~193 files to ~65.
    """
    days = list(sessions[::every])
    if sessions and sessions[-1] not in days:
        days.append(sessions[-1])
    return tuple(days)


# -- excel --------------------------------------------------------------------

def to_excel(sheets, cover_rows):
    """
    Branded workbook. The cover sheet carries every run parameter, so a file that
    has travelled by email still says what produced it.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        pd.DataFrame(cover_rows, columns=["Parameter", "Value"]).to_excel(
            w, sheet_name="Cover", index=False)
        for name, df in sheets:
            df.to_excel(w, sheet_name=name[:31], index=False)

        wb = w.book
        head_fill = PatternFill("solid", fgColor=P["primary"].lstrip("#"))
        head_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        body_font = Font(name="Calibri", size=10)
        band = PatternFill("solid", fgColor=P["tint8"].lstrip("#"))
        thin = Side(style="thin", color=P["rule"].lstrip("#"))
        pos_f = Font(name="Calibri", size=10, color=P["pos"].lstrip("#"))
        neg_f = Font(name="Calibri", size=10, color=P["neg"].lstrip("#"))
        signed = {"outperf pp", "daypl", "dayreturnpct", "dayreturnpct_tr", "total",
                  "realised", "unrealised", "contributionpp", "driftpp",
                  "deviationpp", "portfolio %", "drift pp"}

        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue
            headers = [c.value for c in ws[1]]
            for c in ws[1]:
                c.fill, c.font = head_fill, head_font
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(bottom=thin)
            ws.row_dimensions[1].height = 20
            ws.freeze_panes = "A2"

            for j, h in enumerate(headers, start=1):
                key = str(h or "").strip().lower()
                is_num = False
                width = max(11, min(30, len(str(h or "")) + 4))
                for i in range(2, min(ws.max_row, 400) + 1):
                    cell = ws.cell(row=i, column=j)
                    cell.font = body_font
                    v = cell.value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        is_num = True
                        cell.number_format = "#,##0.00" if abs(v) < 1e5 else "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                        if key in signed:
                            cell.font = pos_f if v >= 0 else neg_f
                    else:
                        width = max(width, min(46, len(str(v or "")) + 3))
                if is_num:
                    width = max(width, 13)
                ws.column_dimensions[get_column_letter(j)].width = width

            for i in range(2, ws.max_row + 1):
                if i % 2 == 0:
                    for j in range(1, ws.max_column + 1):
                        ws.cell(row=i, column=j).fill = band
    return buf.getvalue()


# -- sidebar ------------------------------------------------------------------

with st.sidebar:
    st.markdown("### Portfolio")
    portfolio_name = st.text_input(
        "Name", value="", placeholder="e.g. MO Technical Focus",
        help="Used in the page header, the factsheet and the export filenames. "
             "Falls back to the uploaded file name, or \"pasted log\", when left blank.")
    capital_label = st.selectbox("Starting capital", list(CAPITAL_OPTIONS),
                                 index=list(CAPITAL_OPTIONS).index(DEFAULT_CAPITAL))
    capital = CAPITAL_OPTIONS[capital_label]

    start_override = st.date_input(
        "Start Date", value=date(2025, 10, 20),
        min_value=date(2015, 1, 1), max_value=date.today(),
        help="Day 1 of the NAV series - the date the client's money goes in. "
             "The portfolio opens with the model's allocation as it stood on this "
             "date; earlier advice is not replayed as trades. Those opening "
             "positions fill at this date's closing price, since a client starting "
             "today cannot buy at a price quoted weeks ago.")
    to_date = st.date_input("Calculate through", value=date.today(),
                            min_value=date(2015, 1, 1), max_value=date.today())

    st.markdown("### Rebalancing")
    gate_pp = st.number_input(
        "Rebalance Tolerance (percentage points)", 0.0, 10.0, 1.0, 0.25,
        help="How far a holding may drift from its model weight before it is "
             "traded back - in either direction. At 1pp, a 10% model position is "
             "left alone anywhere between 9% and 11%. Outside that band it is "
             "traded back to exactly 10%: trimmed if it has run up, topped up if "
             "it has fallen behind.")
    force_retarget = st.checkbox(
        "Rebalance all holdings on advice dates", value=True,
        help="On: whenever any advice is logged, every holding is checked against "
             "its model weight, and any that has drifted beyond the tolerance is "
             "traded back - even holdings that advice did not mention. Off: only "
             "the symbols named in that day's advice can trade, and everything "
             "else is left to drift until its own advice arrives.")

    st.markdown("### Income")
    div_on = st.checkbox(
        "Include dividends", value=True,
        help="A stock's price falls by roughly the dividend on its ex-date, and "
             "the exchange close already reflects that. This credits the cash the "
             "client actually receives, so the two offset and the income is not "
             "lost. Trade prices and quantities are never affected - results are "
             "always shown on both bases.")

    st.markdown("### Corporate actions")
    ca_on = st.checkbox(
        "Adjust for corporate actions", value=True,
        help="A split, bonus or demerger cuts the share price without the holder "
             "losing anything. Left alone, that drop reads as a market loss - an "
             "unadjusted 10:1 split looks like a 90% collapse. This fetches NSE's "
             "official corporate-action calendar and restates the affected prices, "
             "in the log as well as the price history.")
    ca_mode = st.radio(
        "Basis", ["pre_ex_down", "post_ex_up"], index=0,
        format_func=lambda m: ("Restate history (recommended)" if m == "pre_ex_down"
                               else "Keep history, scale forward"),
        help="Both give identical returns; they differ in the share count you end "
             "up holding. Restate history rewrites prices from before the event "
             "onto the post-event scale, so trades after it use real market prices "
             "and the share count matches a real client's. Keep history leaves the "
             "past untouched and inflates later prices instead.")
    dm_policy = st.radio(
        "Demergers", ["require", "gap"], index=0,
        format_func=lambda m: ("Require a factor (recommended)" if m == "require"
                               else "Approximate from the price gap"),
        help="Splits and bonuses state their exact terms in NSE's file, so they "
             "are always handled precisely and never need input. Demergers do not "
             "- how the value divided between the parent and the new company sits "
             "only in the scheme document. Require a factor stops and asks you for "
             "it. Approximate guesses from the size of the price drop, which "
             "assumes the stock had no genuine move that day.")

    file_ov, file_rej = engine.load_override_file()
    if file_ov:
        st.caption("From `ca_overrides.csv`: "
                   + ", ".join(f"{sym} {ex}" for sym, ex in sorted(file_ov)))
    for bad, why in file_rej:
        st.caption(f"`ca_overrides.csv` - {bad}: {why}")

    ca_override_txt = st.text_area(
        "Demerger factors (this run only)", height=76,
        placeholder="TRIVENI, 2026-07-22, -5.39%",
        help="One per line: SYMBOL, EX-DATE, VALUE. VALUE is either the stock's "
             "genuine return on the ex-date written with a % sign, or an explicit "
             "price factor. Entries here override ca_overrides.csv for this run.")
    ca_method = st.radio(
        "Ex-date factor", ["gap", "index"], index=0, horizontal=True,
        format_func=lambda m: "Price gap" if m == "gap" else "Index-relative",
        help="Only used when Demergers is set to Approximate.")

    st.markdown("### Comparison and costs")
    _bench_options = (["None"] + NSE_NATIVE_BENCHMARKS
                      + ["Sensex", "BSE 500"])
    benchmark = st.selectbox(
        "Benchmark", _bench_options,
        index=_bench_options.index("Nifty 500"),
        help="Price return, matching the portfolio's price-return basis. A price "
             "index excludes its own constituents' dividends, so outperformance "
             "measured on the dividend-inclusive basis is flattered by roughly "
             "the index yield. Nifty-family indices are fetched from NSE's own "
             "archive; Sensex and BSE 500 use a yfinance fallback since NSE "
             "does not publish BSE indices.")
    brokerage_pct = st.number_input(
        "Brokerage (% per side)", 0.00, 5.00, 0.00, 0.01, format="%.2f",
        help="Enter as a percentage, e.g. 0.20 for 0.20%. Applied to two-sided "
             "turnover as a reporting overlay. It never changes a trade quantity, "
             "so gross and net stay comparable.")
    brokerage_bps = brokerage_pct * 100.0
    statutory_pct = st.number_input(
        "Statutory charges (% per side)", 0.00, 5.00, 0.00, 0.01, format="%.2f",
        help="STT, exchange charges, stamp duty and GST as one combined rate. "
             "Applied to the same two-sided turnover as brokerage, and to both "
             "the cash and MTF models identically.")

    with st.expander("2x MTF leverage comparison"):
        mtf_on = st.checkbox(
            "Compare against a 2x MTF version", value=False,
            help="Runs the identical strategy with every stock purchase funded "
                 "50% by the client's own money and 50% by MTF borrowing. Adds "
                 "a comparison tab. Does not change the NAV figures above it.")
        mtf_rate = st.number_input(
            "MTF interest (% p.a.)", 0.0, 60.0, 15.0, 0.25, format="%.2f",
            help="Simple interest on the outstanding borrowing, charged on every "
                 "calendar day including weekends. Placeholder until the broker's "
                 "actual rate is confirmed.")
        mtf_pledge = st.number_input(
            "Pledge / unpledge per trade leg (Rs)", 0.0, 500.0, 29.50, 0.50,
            format="%.2f",
            help="Charged once per buy and once per sell. Rs 25 plus 18% GST by "
                 "default. Placeholder pending the broker's fee schedule.")
        st.caption("Funding split is fixed at 50% own money / 50% borrowed at "
                   "every purchase. Brokerage and statutory rates are shared "
                   "with the cash model above, so the two are comparable.")

    st.markdown("### Advice log")
    src = st.radio("Source", ["Upload file", "Paste rows"], horizontal=True,
                   label_visibility="collapsed", key="log_src")
    uploaded, pasted = None, ""
    if src == "Upload file":
        uploaded = st.file_uploader("Upload", type=["csv", "xlsx", "xls"],
                                    label_visibility="collapsed",
                                    help="CSV or XLSX. Delimiter and columns "
                                         "auto-detected.")
    else:
        pasted = st.text_area(
            "Paste", height=150, label_visibility="collapsed",
            placeholder="Paste rows copied from Excel, including the header row.",
            help="Copy the log rows in Excel and paste here. Excel copies as "
                 "tab-separated text, which the parser already detects.")
        if pasted.strip():
            st.caption(f"{len(pasted.strip().splitlines()) - 1} data row(s) detected")

    log_source = uploaded
    if src == "Paste rows" and pasted.strip():
        log_source = _PastedLog(pasted.encode("utf-8"))

    run_btn = st.button("Calculate NAV", type="primary", use_container_width=True,
                        disabled=log_source is None)
    st.caption("IMP NAV Calculator v3.1")

# -- landing ------------------------------------------------------------------

if log_source is None:
    st.markdown('<div class="rh-title">NAV Calculator</div>', unsafe_allow_html=True)
    st.markdown('<div class="rh-strip"><div class="rh-item"><span class="rh-k">'
                'Model portfolio simulation</span><span class="rh-v">Daily rebased '
                'NAV &middot; target-weight rebalancing &middot; official NSE closing '
                'prices</span></div></div>', unsafe_allow_html=True)
    st.markdown('<div class="note"><strong>Upload or paste an advice log to begin.'
                '</strong><br>CSV or XLSX (upload), or tab-separated rows copied '
                'from Excel (paste). Delimiter and column structure are detected '
                'automatically.</div>', unsafe_allow_html=True)
    with st.expander("How it works"):
        st.markdown("""
**Model, not replay.** Each log row states the model's target weight for a symbol.
A trade happens only when the client portfolio has drifted more than the tolerance
away from that target - it is not triggered by the row existing. Roughly a third of
the rows in a typical log are weight-drift entries the dashboard writes as prices
move; the tolerance filters them out by construction.

**Sequence on any date carrying log rows**
1. Same-date rows per symbol are chain-resolved along the `old -> new` weight trail
   (never by `No.`) and collapse to the last link. Its price is the execution price:
   `Entry Price` when the old weight is 0, otherwise `Modified Price`.
2. Model target weights update.
3. Every holding is measured against its target using the **previous** EOD NAV and
   previous EOD closes.
4. Where the deviation exceeds the tolerance, quantity is floored to the target
   weight. Sells execute before buys.
5. Day-end NAV = holdings at today's EOD closes + cash.

**Cash** is idle allocation plus any LIQUIDCASE / liquid-ETF weight, held as one
pool earning nothing. Negative cash is permitted and always alerted.

**Dividends** are credited to cash on the ex-date; trade prices are never
adjusted. Results are shown both including and excluding dividends throughout.

**Corporate actions** are read from NSE's official calendar. Splits and bonuses are
exact; demergers require a factor you supply, because the terms sit only in the
scheme document - either in the sidebar for a single run, or in `ca_overrides.csv`
at the repo root so it applies to every future run.
        """)
    st.stop()

# -- parse --------------------------------------------------------------------

log_df, parse_err = engine.load_log(log_source)
if parse_err:
    st.markdown(f'<div class="err">{parse_err}</div>', unsafe_allow_html=True)
    st.stop()

pname = portfolio_name.strip() or log_source.name.rsplit(".", 1)[0].replace("_", " ")
eq_log = log_df[~log_df["is_liquid"]]

# Cache-key caution: hash the paste, never the raw string - log_source.name is
# a constant for every paste ("pasted_log.csv"), so it can't distinguish two
# different pastes, and re-hashing a large paste on every widget change is
# wasted work. Uploads keep the filename; it already changes with the file.
log_key = uploaded.name if src == "Upload file" else hashlib.md5(pasted.encode()).hexdigest()

params = (log_key, capital, start_override, to_date, gate_pp, force_retarget,
          benchmark, ca_on, ca_mode, ca_method, dm_policy, ca_override_txt,
          div_on, mtf_on)
if run_btn:
    st.session_state["nav_run"] = True
if st.session_state.get("nav_params") != params:
    for k in ("nav_results", "nav_excel", "nav_csv", "nav_pdf"):
        st.session_state.pop(k, None)
    st.session_state["nav_params"] = params

if not st.session_state.get("nav_run"):
    st.markdown(f'<div class="rh-title">{pname}</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="rh-strip">'
        f'<div class="rh-item"><span class="rh-k">Log rows</span>'
        f'<span class="rh-v">{len(log_df)}</span></div>'
        f'<div class="rh-item"><span class="rh-k">Symbols</span>'
        f'<span class="rh-v">{eq_log["symbol"].nunique()}</span></div>'
        f'<div class="rh-item"><span class="rh-k">Advice dates</span>'
        f'<span class="rh-v">{eq_log["mod_dt"].nunique()}</span></div>'
        f'<div class="rh-item"><span class="rh-k">Log period</span>'
        f'<span class="rh-v">{eq_log["mod_dt"].min():%d %b %Y} &ndash; '
        f'{eq_log["mod_dt"].max():%d %b %Y}</span></div></div>',
        unsafe_allow_html=True)
    st.markdown('<div class="note">Log parsed. Set the options in the sidebar, then '
                'click <strong>Calculate NAV</strong>.</div>', unsafe_allow_html=True)
    st.stop()

# -- compute (cached so the download buttons cannot wipe it) ------------------

if "nav_results" not in st.session_state:
    px_start = min(eq_log["mod_dt"].min(), start_override) - timedelta(days=7)
    closes, fetch_errors = fetch_closes(px_start, to_date)
    if not closes:
        st.markdown('<div class="err">No NSE sessions retrieved for this range. '
                    'Check the date window and the archive host.</div>',
                    unsafe_allow_html=True)
        st.stop()
    bench = benchmark_series(closes, benchmark, start_override, to_date)

    ca_overrides, ca_bad = engine.load_override_file()
    typed, typed_bad = engine.parse_override_lines(ca_override_txt)
    ca_overrides.update(typed)
    ca_bad += typed_bad

    ca_alerts, ca_actions, dividends, div_paid, div_dup_suppressed = [], [], {}, [], 0
    for bad, why in ca_bad:
        ca_alerts.append(dict(Date=None, Symbol="-", Type="CA OVERRIDE IGNORED",
                              Detail=f"{bad} - {why}"))

    held = set(log_df.loc[~log_df["is_liquid"], "symbol"])
    if ca_on or div_on:
        with st.spinner("Fetching NSE corporate-action calendar..."):
            ca_actions = [a for a in fetch_ca(ca_sample_days(sorted(closes)))
                          if a["Symbol"] in held]
    if ca_on:
        closes, log_df, applied = engine.apply_corporate_actions(
            closes, log_df, ca_actions, mode=ca_mode, method=ca_method, bench=bench,
            overrides=ca_overrides, demerger_policy=dm_policy)
        ca_alerts += applied
    if div_on:
        dividends, div_alerts, div_dup_suppressed = engine.build_dividends(ca_actions, held)
        ca_alerts += div_alerts
    ca_alerts += engine.gap_detector(closes, log_df, ca_actions, threshold_pct=20.0)

    with st.spinner("Running the NAV engine..."):
        try:
            nav_rows, trades, alerts, div_paid = engine.run_nav(
                log_df, closes, capital, start_override, to_date,
                gate_pp=gate_pp, force_retarget=force_retarget, dividends=dividends)
        except ValueError as e:
            st.markdown(f'<div class="err">{e}</div>', unsafe_allow_html=True)
            st.stop()

    # -- 2x MTF overlay ------------------------------------------------
    # The sizing run below exists ONLY to produce trade quantities at ~2x
    # buying power, so that share flooring, the cash cap and corporate
    # actions are all handled by one engine rather than a parallel copy.
    # Its own NAV series is NOT a result and is never stored or displayed.
    mtf_rows = None
    if mtf_on:
        with st.spinner("Running the 2x MTF comparison..."):
            try:
                _sz_nav, sz_trades, _sz_alerts, _sz_div = engine.run_nav(
                    log_df, closes, capital * 2.0, start_override, to_date,
                    gate_pp=gate_pp, force_retarget=force_retarget,
                    dividends=dividends)
                _cal = [r["Date"] for r in nav_rows]
                mtf_rows = mtf.run_leverage(sz_trades, closes, _cal, capital)
                _last = {}
                for _d in _cal:
                    for _s, _p in closes.get(_d, {}).items():
                        if _p and _p > 0:
                            _last[_s] = _p
                _re, _un = mtf.fifo_pnl(sz_trades, _last)
                _gap = mtf_rows[-1]["NetWorth"] - (capital + _re + _un)
                mtf_audit = dict(realised=_re, unrealised=_un, gap=_gap,
                                 turnover=sum(t["Value"] for t in sz_trades),
                                 trades=len(sz_trades))
            except Exception as e:
                mtf_rows = None
                alerts.append(dict(Date=to_date, Symbol="-",
                                   Type="MTF COMPARISON FAILED",
                                   Detail=f"{e} - cash model figures are unaffected."))
        if mtf_rows is not None:
            st.session_state["mtf_results"] = (mtf_rows, mtf_audit)
        else:
            st.session_state.pop("mtf_results", None)
    else:
        st.session_state.pop("mtf_results", None)

    for d, err in sorted(fetch_errors.items()):
        alerts.append(dict(Date=d, Symbol="-", Type="BHAVCOPY FETCH ERROR", Detail=err))
    st.session_state["nav_results"] = (nav_rows, trades, ca_alerts + alerts,
                                       bench, div_paid, div_dup_suppressed)

nav_rows, trades, alerts, bench, div_paid, div_dup_suppressed = st.session_state["nav_results"]

nav_df = pd.DataFrame([{k: v for k, v in r.items() if k != "_holdings"}
                       for r in nav_rows])
trades_df = pd.DataFrame(trades) if trades else pd.DataFrame(
    columns=["Date", "Symbol", "Side", "Qty", "Price", "Value"])
alerts_df = pd.DataFrame(alerts) if alerts else pd.DataFrame(
    columns=["Date", "Symbol", "Type", "Detail"])
recon_df = engine.reconciliation(nav_rows)
wt_matrix = engine.holdings_matrix(nav_rows, "weight")
qty_matrix = engine.holdings_matrix(nav_rows, "qty")
bname = benchmark if benchmark != "None" else "Benchmark"
perf_df = engine.period_returns(nav_rows, capital, bench, bname)
perf_tr_df = engine.period_returns(nav_rows, capital, bench, bname, key="NAV_TR",
                                   div_events=div_paid)
attrib_df = engine.attribution(nav_rows, trades, div_paid)
div_df = pd.DataFrame(div_paid) if div_paid else pd.DataFrame(
    columns=["Date", "Symbol", "DPS", "Qty", "Amount"])
risk = engine.risk_stats(nav_rows, capital)
costs = engine.cost_impact(trades, nav_rows, capital, brokerage_bps)
cashd = engine.cash_drag(nav_rows)

last, first = nav_rows[-1], nav_rows[0]
total_return = (last["NAV"] - capital) / capital * 100
total_return_tr = (last["NAV_TR"] - capital) / capital * 100
abs_pl = last["NAV"] - capital
div_total = last["DividendCash"]
_since = perf_df[perf_df["Period"] == "Since launch"]
outperf = _since.iloc[0]["Outperf pp"] if len(_since) else None
outperf = None if outperf is None or outperf != outperf else float(outperf)

# -- report header ------------------------------------------------------------

st.markdown(f'<div class="rh-title">{pname}</div>', unsafe_allow_html=True)
st.markdown(
    f'<div class="rh-strip">'
    f'<div class="rh-item"><span class="rh-k">Period</span><span class="rh-v">'
    f'{first["Date"]:%d %b %Y} &ndash; {last["Date"]:%d %b %Y}</span></div>'
    f'<div class="rh-item"><span class="rh-k">Sessions</span>'
    f'<span class="rh-v">{len(nav_rows)}</span></div>'
    f'<div class="rh-item"><span class="rh-k">Capital</span>'
    f'<span class="rh-v">{capital_label}</span></div>'
    f'<div class="rh-item"><span class="rh-k">Benchmark</span>'
    f'<span class="rh-v">{bname}</span></div>'
    f'<div class="rh-item"><span class="rh-k">Tolerance</span>'
    f'<span class="rh-v">{gate_pp:.2f} pp</span></div>'
    f'<div class="rh-item"><span class="rh-k">Basis</span>'
    f'<span class="rh-v">Price return</span></div>'
    f'</div>', unsafe_allow_html=True)

# -- focal metric + supporting KPIs -------------------------------------------

fc, kc = st.columns([1.05, 2.6], gap="large")
with fc:
    sub = (f"vs {bname} {outperf:+.2f} pp" if outperf is not None
           else "no benchmark selected")
    st.markdown(
        f'<div><span class="focal-val {tone(total_return)}">'
        f'{fmt_pct(total_return)}</span></div>'
        f'<div class="focal-lbl">Return since launch</div>'
        f'<div class="focal-sub">{sub}</div>', unsafe_allow_html=True)

with kc:
    kcols = st.columns(4, gap="small")
    cards = [
        ("Portfolio NAV", fmt_inr(last["NAV"]), "neutral"),
        ("Absolute P&L", fmt_inr(abs_pl), tone(abs_pl)),
        ("Incl. dividends", fmt_pct(total_return_tr), tone(total_return_tr)),
        ("Max drawdown", f"{risk.get('Max drawdown %', 0):.2f}%", "loss"),
    ]
    for col, (lbl, val, cls) in zip(kcols, cards):
        col.markdown(f'<div class="kpi"><div class="kpi-val {cls}">{val}</div>'
                     f'<div class="kpi-lbl">{lbl}</div></div>',
                     unsafe_allow_html=True)

    # Alert types that can plausibly move NAV/return vs types that are purely an
    # audit trail (the engine checked something and took no action). Kept as an
    # explicit allowlist so a new alert type defaults to "needs attention" rather
    # than silently disappearing into the audit tier if this list is not updated.
    IMPACT_ALERTS = {
        "CHAIN UNRESOLVED", "FORCED REBALANCE (no log row)", "LOG PRICE ZERO",
        "NO PRICE - TRADE SKIPPED", "NEGATIVE CASH",
        "BUY UNDERFILLED - INSUFFICIENT CASH", "NO EOD CLOSE",
        "PRE-START ADVICE - REPRICED", "ADVICE DATE ROLLED",
        "ADVICE DATE ORPHANED", "UNEXPLAINED PRICE GAP",
        "CORPORATE ACTION - CANNOT ADJUST", "CORPORATE ACTION - FACTOR REQUIRED",
        "DIVIDEND NOT PARSED",
        "CA OVERRIDE IGNORED", "BHAVCOPY FETCH ERROR",
    }
    n_impact = int(alerts_df["Type"].isin(IMPACT_ALERTS).sum()) if len(alerts_df) else 0

    if n_impact:
        st.markdown(f'<div style="margin-top:12px"><span class="chip chip-warn">'
                    f'<span class="dot"></span>{n_impact} alert(s) need '
                    f'attention</span></div>', unsafe_allow_html=True)
    elif len(alerts_df):
        st.markdown(f'<div style="margin-top:12px"><span class="chip chip-ok">'
                    f'<span class="dot"></span>No alerts need attention '
                    f'({len(alerts_df)} audit-only entries below)</span></div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div style="margin-top:12px"><span class="chip chip-ok">'
                    '<span class="dot"></span>No alerts</span></div>',
                    unsafe_allow_html=True)

if len(alerts_df):
    impact_df = alerts_df[alerts_df["Type"].isin(IMPACT_ALERTS)]
    audit_df = alerts_df[~alerts_df["Type"].isin(IMPACT_ALERTS)]

    with st.expander(f"Needs attention ({len(impact_df)})",
                     expanded=len(impact_df) > 0):
        if len(impact_df):
            types = sorted(impact_df["Type"].unique())
            pick = st.multiselect("Filter by type", types, default=types,
                                  key="alert_filter_impact")
            shown = impact_df[impact_df["Type"].isin(pick)] if pick else impact_df
            st.caption(f"Showing {len(shown)} of {len(impact_df)}")
            st.dataframe(shown, use_container_width=True, hide_index=True,
                        height=300)
        else:
            st.caption("None.")

    with st.expander(f"Audit trail — no return impact ({len(audit_df)})",
                     expanded=False):
        st.caption("The engine checked these and confirmed no action was needed "
                   "— e.g. a corporate action was found but the symbol was not "
                   "held on the ex-date, or a demerger factor was required but "
                   "the position was never held across the ex-date. Kept for "
                   "verification; none of these change NAV or return.")
        if len(audit_df):
            atypes = sorted(audit_df["Type"].unique())
            apick = st.multiselect("Filter by type", atypes, default=atypes,
                                   key="alert_filter_audit")
            ashown = audit_df[audit_df["Type"].isin(apick)] if apick else audit_df
            st.caption(f"Showing {len(ashown)} of {len(audit_df)}")
            st.dataframe(ashown, use_container_width=True, hide_index=True,
                        height=300)
        else:
            st.caption("None.")

# -- chart --------------------------------------------------------------------

st.markdown('<div class="sec">Rebased NAV &mdash; base 100 at inception</div>',
            unsafe_allow_html=True)

dates_plot = [r["Date"] for r in nav_rows]
rebased = [r["Rebased"] for r in nav_rows]
rebased_tr = [r["Rebased_TR"] for r in nav_rows]

# Running peak drives the drawdown shading - the visual answer to "how bad did
# this get", which a bare NAV line never shows.
_peak, peak_line = -1e18, []
for v in rebased:
    _peak = max(_peak, v)
    peak_line.append(_peak)

fig = go.Figure()
fig.add_trace(go.Scatter(x=dates_plot, y=peak_line, name="peak",
                         line=dict(width=0), hoverinfo="skip", showlegend=False))
fig.add_trace(go.Scatter(x=dates_plot, y=rebased, name="Portfolio", fill="tonexty",
                         fillcolor="rgba(192,57,43,.06)",
                         line=dict(color=P["primary"], width=2.2),
                         hovertemplate="<b>%{x|%d %b %Y}</b><br>NAV %{y:.2f}"
                                       "<br>Return %{customdata:+.2f}%<extra></extra>",
                         customdata=[v - 100 for v in rebased]))
if div_total > 0:
    fig.add_trace(go.Scatter(x=dates_plot, y=rebased_tr, name="Incl. dividends",
                             line=dict(color=P["pos"], width=1.2, dash="dot"),
                             hovertemplate="incl. div %{y:.2f}<extra></extra>"))
if bench:
    bd = sorted(d for d in bench if first["Date"] <= d <= last["Date"])
    if bd:
        base = bench[bd[0]]
        fig.add_trace(go.Scatter(x=bd, y=[bench[d] / base * 100 for d in bd],
                                 name=bname,
                                 line=dict(color=P["muted"], width=1.3, dash="dash"),
                                 hovertemplate=f"{bname} %{{y:.2f}}<extra></extra>"))

fig.add_hline(y=100, line_dash="dot", line_color=P["rule"], line_width=1)
fig.add_annotation(x=dates_plot[-1], y=rebased[-1], text=f"  {rebased[-1]:.1f}",
                   showarrow=False, xanchor="left",
                   font=dict(size=12, color=P["primary"],
                             family="Plus Jakarta Sans"))
fig.update_layout(
    height=380, hovermode="x unified",
    legend=dict(orientation="h", yanchor="bottom", y=1.0, xanchor="left", x=0,
                font=dict(size=11, color=P["muted"]), bgcolor="rgba(0,0,0,0)"),
    xaxis=dict(showgrid=False, tickformat="%b %y", showline=True,
               linecolor=P["rule"], ticks="outside", tickcolor=P["rule"],
               tickfont=dict(size=11, color=P["muted"])),
    yaxis=dict(title="", gridcolor=P["rule"], griddash="dot", zeroline=False,
               showline=False, tickfont=dict(size=11, color=P["muted"])),
    plot_bgcolor="white", paper_bgcolor="white",
    font=dict(family="Inter, sans-serif", color=P["ink"]),
    margin=dict(l=8, r=52, t=34, b=8))
st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
st.caption("Shaded area is drawdown from the running peak. Dividends are excluded "
           "from the primary NAV line.")

# -- performance vs benchmark -------------------------------------------------

st.markdown('<div class="sec">Performance vs benchmark</div>', unsafe_allow_html=True)
bcol = f"{bname} %"
basis = st.radio("Basis", ["Excluding dividends", "Including dividends"],
                 horizontal=True, key="perf_basis", label_visibility="collapsed")
_pf = perf_df if basis == "Excluding dividends" else perf_tr_df
st.dataframe(
    _pf, use_container_width=True, hide_index=True,
    column_config={
        "Days": st.column_config.NumberColumn(format="%d"),
        "Portfolio %": st.column_config.NumberColumn(format="%.2f"),
        bcol: st.column_config.NumberColumn(format="%.2f"),
        "Outperf pp": st.column_config.NumberColumn(format="%.2f"),
    })
cap = (f"Point-to-point price returns on both legs. Base session is the last session "
       f"on or before each period anchor; periods starting before inception show n/a. "
       f"Cash weight on {last['Date']:%d %b %Y} is "
       f"{(last['Cash'] / last['NAV'] * 100 if last['NAV'] else 0):.1f}% &mdash; the "
       f"benchmark is fully invested, so part of any gap is cash drag.")
if basis == "Including dividends":
    cap = (f"Dividends received {fmt_inr(div_total)} "
           f"({total_return_tr - total_return:+.2f} pp). The benchmark remains a price "
           f"index and excludes its constituents' dividends, so outperformance on this "
           f"basis is flattered by roughly the index yield. ") + cap
st.caption(cap)

# -- holdings + summary -------------------------------------------------------

holdings_export = pd.DataFrame([
    {"Symbol": s, "Qty": h["qty"], "Close": round(h["price"], 2),
     "Value": round(h["value"], 2),
     "Achieved Wt %": round(h["value"] / last["NAV"] * 100, 2) if last["NAV"] else 0.0,
     "Model Wt %": round(h["model_wt"], 2),
     "Drift pp": round((h["value"] / last["NAV"] * 100 if last["NAV"] else 0.0)
                       - h["model_wt"], 2)}
    for s, h in sorted(last["_holdings"].items(), key=lambda x: -x[1]["value"])
])

col_h, col_s = st.columns([3, 2], gap="large")
with col_h:
    st.markdown(f'<div class="sec">Holdings &mdash; {last["Date"]:%d %b %Y}</div>',
                unsafe_allow_html=True)
    if len(holdings_export):
        st.dataframe(holdings_export, use_container_width=True, hide_index=True,
                     height=min(430, 42 + 35 * len(holdings_export)))
        st.caption(f"Largest drift vs model "
                   f"{holdings_export['Drift pp'].abs().max():.2f} pp. Drift inside "
                   f"the {gate_pp:.2f} pp tolerance is expected and untraded.")
    else:
        st.info("No open positions on this date.")

with col_s:
    st.markdown('<div class="sec">Summary</div>', unsafe_allow_html=True)
    buys = int((trades_df["Side"] == "BUY").sum()) if len(trades_df) else 0
    sells = int((trades_df["Side"] == "SELL").sum()) if len(trades_df) else 0
    turnover = float(trades_df["Value"].sum()) if len(trades_df) else 0.0
    summary = {
        "Market value": fmt_inr(last["MarketValue"]),
        "Cash": f'{fmt_inr(last["Cash"])} '
                f'({(last["Cash"] / last["NAV"] * 100) if last["NAV"] else 0:.1f}%)',
        "Return excl. dividends": fmt_pct(total_return),
        "Return incl. dividends": fmt_pct(total_return_tr),
        "Dividends received": fmt_inr(div_total),
        "Annualised volatility": f"{risk.get('Annualised volatility %', 0):.2f}%",
        "Positive days": f"{risk.get('Positive days %', 0)}%",
        "Days to recover drawdown": str(risk.get("Days to recover", "-")),
        "Event dates": str(sum(1 for r in nav_rows if r["Type"] == "TRADE")),
        "Trades (buy / sell)": f"{buys} / {sells}",
        "Two-sided turnover": f"{fmt_inr(turnover)} "
                              f"({turnover / capital * 100:.0f}% of capital)",
        "Open positions": str(len(last["_holdings"])),
        "Min cash over period": fmt_inr(nav_df["Cash"].min()),
    }
    st.dataframe(pd.DataFrame(list(summary.items()), columns=["Metric", "Value"]),
                 use_container_width=True, hide_index=True, height=490)

# -- detail tabs --------------------------------------------------------------

st.markdown('<div class="sec">Detail</div>', unsafe_allow_html=True)
_mtf_state = st.session_state.get("mtf_results")
_names = ["Contributors", "Risk & costs", "Daily NAV", "Holdings on a date",
          "Holdings matrix", "Trades", "Dividends", "Reconciliation"]
if _mtf_state:
    _names.insert(2, "2x MTF comparison")
_tabs = st.tabs(_names)
if _mtf_state:
    (tab_attr, tab_risk, tab_mtf, tab_nav, tab_book, tab_matrix,
     tab_trades, tab_div, tab_recon) = _tabs
else:
    tab_mtf = None
    (tab_attr, tab_risk, tab_nav, tab_book, tab_matrix,
     tab_trades, tab_div, tab_recon) = _tabs

with tab_attr:
    if len(attrib_df):
        _cols = ["Symbol", "AbsReturnPct", "ContributionPP", "Total",
                "Realised", "Unrealised", "Dividends", "Invested", "StillHeld"]
        _view = attrib_df[[c for c in _cols if c in attrib_df.columns]]
        _cfg = {
            "AbsReturnPct": st.column_config.NumberColumn(
                "Position return %", format="%.2f",
                help="This symbol's own P&L as a % of capital deployed to it."),
            "ContributionPP": st.column_config.NumberColumn(
                "Contribution pp", format="%.2f",
                help="This symbol's rupee P&L as a % of the whole portfolio's "
                     "starting capital."),
        }
        ca_, cb_ = st.columns(2, gap="large")
        with ca_:
            st.markdown("**Key contributors**")
            st.dataframe(_view.head(10), use_container_width=True,
                         hide_index=True, height=380, column_config=_cfg)
        with cb_:
            st.markdown("**Key detractors**")
            st.dataframe(_view.tail(10).iloc[::-1], use_container_width=True,
                         hide_index=True, height=380, column_config=_cfg)
        st.caption("Realised P&L uses FIFO lots, so a symbol traded in and out "
                   "repeatedly is not flattered by averaging across separate holding "
                   "periods. Position return % is this symbol's own gain/loss on "
                   "capital deployed to it; ContributionPP is the rupee total as a "
                   "percentage of starting capital; the column sums to total P&L "
                   "including "
                   "dividends.")
    else:
        st.info("No closed or open positions to attribute.")

with tab_risk:
    r1, r2, r3 = st.columns(3, gap="large")
    with r1:
        st.markdown("**Risk**")
        st.dataframe(pd.DataFrame(list(risk.items()), columns=["Metric", "Value"]),
                     use_container_width=True, hide_index=True, height=420)
    with r2:
        st.markdown("**Turnover & cost**")
        st.dataframe(pd.DataFrame(list(costs.items()), columns=["Metric", "Value"]),
                     use_container_width=True, hide_index=True, height=300)
        if brokerage_bps == 0:
            st.caption("Set a brokerage rate in the sidebar to see the cost drag.")
    with r3:
        st.markdown("**Cash**")
        st.dataframe(pd.DataFrame(list(cashd.items()), columns=["Metric", "Value"]),
                     use_container_width=True, hide_index=True, height=300)
        st.caption("Cash earns nothing in this model, so a high average cash weight "
                   "is a real drag relative to a fully invested index.")

with tab_nav:
    flt = st.selectbox("Rows", ["All", "Event days only", "MTM days only"],
                       key="navflt")
    dsel = nav_df.copy()
    if flt == "Event days only":
        dsel = dsel[dsel["Type"] == "TRADE"]
    elif flt == "MTM days only":
        dsel = dsel[dsel["Type"] == "MTM"]
    st.dataframe(dsel.iloc[::-1].reset_index(drop=True), use_container_width=True,
                 hide_index=True, height=420,
                 column_config={
                     "Rebased": st.column_config.NumberColumn(format="%.4f"),
                     "DayReturnPct": st.column_config.NumberColumn(format="%.3f"),
                     "MarketValue": st.column_config.NumberColumn(format="%.2f"),
                     "Cash": st.column_config.NumberColumn(format="%.2f"),
                     "NAV": st.column_config.NumberColumn(format="%.2f"),
                     "DayPL": st.column_config.NumberColumn(format="%.2f"),
                     "NAV_TR": st.column_config.NumberColumn(format="%.2f"),
                     "Rebased_TR": st.column_config.NumberColumn(format="%.4f"),
                     "DividendCash": st.column_config.NumberColumn(format="%.2f"),
                 })
    st.caption("NAV excludes dividends; NAV_TR includes them. DividendCash is the "
               "running total received.")

with tab_book:
    sessions = [r["Date"] for r in nav_rows]
    pick = st.selectbox("Session", sessions[::-1], index=0, key="book_date",
                        format_func=lambda d: d.strftime("%d %b %Y (%a)"))
    row, book = engine.holdings_on(nav_rows, pick)
    if row is not None:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("NAV", fmt_inr(row["NAV"]))
        m2.metric("Market value", fmt_inr(row["MarketValue"]))
        m3.metric("Cash", fmt_inr(row["Cash"]),
                  f"{row['Cash'] / row['NAV'] * 100:.1f}% of NAV"
                  if row["NAV"] else None)
        m4.metric("Positions", str(row["Positions"]))
        if len(book):
            st.dataframe(book, use_container_width=True, hide_index=True, height=420)
            st.caption(f"{row['Type']} day. DriftPP inside the {gate_pp:.2f} pp "
                       "tolerance is expected and untraded.")
        else:
            st.info("Fully in cash on this session.")

with tab_matrix:
    measure = st.radio("Measure", ["Weight %", "Quantity"], horizontal=True, key="mx")
    mx = wt_matrix if measure == "Weight %" else qty_matrix
    st.dataframe(mx.iloc[::-1].reset_index(drop=True), use_container_width=True,
                 hide_index=True, height=460)
    st.caption("TOTAL is holdings + cash and should read 100.00 on every row "
               "(rounding only). Any other value is a bug - report it."
               if measure == "Weight %" else
               "Share counts held at each session close. Blank = not held.")

with tab_trades:
    if len(trades_df):
        st.dataframe(trades_df.iloc[::-1].reset_index(drop=True),
                     use_container_width=True, hide_index=True, height=420)
        st.caption("PriceSource: log = from the advice log | eod-forced = rebalance "
                   "with no log row that date | eod-zerofill = log price was 0 | "
                   "start-close = advice predated the start date.")
    else:
        st.info("No trades generated.")

with tab_div:
    if len(div_df):
        st.dataframe(div_df.iloc[::-1].reset_index(drop=True),
                     use_container_width=True, hide_index=True, height=380)
        st.caption(f"Total {fmt_inr(div_total)} across {len(div_df)} events "
                   f"({total_return_tr - total_return:+.2f} pp). Credited to cash on "
                   "the ex-date; trade prices and quantities are unaffected.")
        if div_dup_suppressed:
            st.caption(f"{div_dup_suppressed} duplicate NSE series listing"
                       f"{'s' if div_dup_suppressed != 1 else ''} removed.")
    elif div_on:
        st.info("No dividends fell on a date when the relevant stock was held.")
    else:
        st.info("Dividends are switched off in the sidebar.")

with tab_recon:
    st.dataframe(recon_df.iloc[::-1].reset_index(drop=True),
                 use_container_width=True, hide_index=True, height=420)
    st.caption("Model weight vs achieved weight per holding per day. "
               "DriftPP inside the tolerance is expected.")


# -- 2x MTF comparison tab ----------------------------------------------------

if tab_mtf is not None:
    with tab_mtf:
        _rows, _audit = st.session_state["mtf_results"]
        _c = mtf.apply_costs(_rows, interest_pa=mtf_rate / 100.0,
                             pledge=mtf_pledge, brokerage_pct=brokerage_pct,
                             statutory_pct=statutory_pct)

        # Cash model, costed on the SAME rates so the two columns are
        # comparable. The headline above this tab is still gross - that is a
        # known open item and is stated in the note below.
        _cash_turnover = sum(t["Value"] for t in trades)
        _cash_cost = _cash_turnover * (brokerage_pct + statutory_pct) / 100.0
        _cash_gross_end = nav_rows[-1]["NAV"]
        _cash_end = _cash_gross_end - _cash_cost
        _cash_ret = _cash_end / capital - 1
        _mtf_end = _c[-1]["NetAfterCosts"]
        _mtf_ret = _mtf_end / capital - 1

        st.markdown(
            '<div class="warn">This comparison does <strong>not</strong> simulate a '
            'margin call or forced selling. In reality the broker would have '
            'stopped new borrowing, and likely forced a sale, well before the '
            'worst points shown here - so real losses in a bad stretch could '
            'have been locked in earlier and at worse prices. The interest rate '
            'and pledge charge are placeholders until the broker\'s actual fee '
            'schedule is confirmed.</div>', unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("Cash model - no borrowing", f"{_cash_ret * 100:.2f}%",
                  help="Return on the client's own money after brokerage and "
                       "statutory charges.")
        c2.metric(f"2x MTF - at {mtf_rate:.2f}% p.a.", f"{_mtf_ret * 100:.2f}%",
                  delta=f"{(_mtf_ret - _cash_ret) * 100:+.2f} pp",
                  help="Return on the same own money after brokerage, statutory "
                       "charges, MTF interest and pledge fees.")
        c3.metric("Extra rupees earned", f"{_mtf_end - _cash_end:,.0f}",
                  help="Difference in the final value of the client's own money.")

        _cash_dd, _, _, _ = mtf.max_drawdown([r["NAV"] for r in nav_rows])
        _mtf_dd, _s_i, _e_i, _pk = mtf.max_drawdown([r["NetWorth"] for r in _rows])
        _ratios = [r["StockPerRupee"] for r in _rows if r["StockPerRupee"]]
        _avg_lev = sum(_ratios) / len(_ratios) if _ratios else 0.0

        st.markdown(
            f'<div class="note">Borrowing turned <strong>{_cash_ret * 100:.2f}%</strong> '
            f'into <strong>{_mtf_ret * 100:.2f}%</strong> on the same '
            f'Rs {capital:,.0f}, and turned a worst fall of '
            f'<strong>{_cash_dd * 100:.2f}%</strong> into '
            f'<strong>{_mtf_dd * 100:.2f}%</strong>. The client paid '
            f'Rs {_c[-1]["AllCosts"] - _cash_cost:,.0f} more in costs to get it.'
            f'</div>', unsafe_allow_html=True)

        # -- cost ladder ---------------------------------------------------
        st.markdown('<div class="sec">Where the money went</div>',
                    unsafe_allow_html=True)
        _ladder = pd.DataFrame([
            ("Client's own money put in", capital, 0.0, capital, 0.0),
            ("Value before any costs", _cash_gross_end,
             (_cash_gross_end / capital - 1) * 100,
             _rows[-1]["NetWorth"], (_rows[-1]["NetWorth"] / capital - 1) * 100),
            ("Less: brokerage", -_cash_turnover * brokerage_pct / 100.0,
             -_cash_turnover * brokerage_pct / 100.0 / capital * 100,
             -_c[-1]["Brokerage"], -_c[-1]["Brokerage"] / capital * 100),
            ("Less: statutory charges", -_cash_turnover * statutory_pct / 100.0,
             -_cash_turnover * statutory_pct / 100.0 / capital * 100,
             -_c[-1]["Statutory"], -_c[-1]["Statutory"] / capital * 100),
            ("Less: MTF interest", 0.0, 0.0,
             -_c[-1]["Interest"], -_c[-1]["Interest"] / capital * 100),
            ("Less: pledge / unpledge", 0.0, 0.0,
             -_c[-1]["Pledge"], -_c[-1]["Pledge"] / capital * 100),
            ("Final value of client's money", _cash_end,
             (_cash_end / capital - 1) * 100, _mtf_end, _mtf_ret * 100),
        ], columns=["Step", "Cash model Rs", "Cash model %",
                    "2x MTF Rs", "2x MTF %"])
        st.dataframe(_ladder, use_container_width=True, hide_index=True,
                     column_config={
                         "Cash model Rs": st.column_config.NumberColumn(format="%.0f"),
                         "2x MTF Rs": st.column_config.NumberColumn(format="%.0f"),
                         "Cash model %": st.column_config.NumberColumn(format="%.2f"),
                         "2x MTF %": st.column_config.NumberColumn(format="%.2f")})
        st.caption(
            f"Every % is measured against the client's own Rs {capital:,.0f}, so the "
            f"two columns compare directly. The MTF book trades about double the "
            f"rupee value (Rs {_audit['turnover']:,.0f} vs Rs {_cash_turnover:,.0f}), "
            f"so its brokerage and statutory charges are about double. That is real, "
            f"not double-counting. Note the headline figures above this tab are still "
            f"gross of costs; only this tab nets them off.")

        # -- chart ---------------------------------------------------------
        _fig = go.Figure()
        _fig.add_trace(go.Scatter(
            x=[r["Date"] for r in nav_rows],
            y=[r["NAV"] / capital * 100 for r in nav_rows],
            name="Cash model", mode="lines", line=dict(width=2)))
        _fig.add_trace(go.Scatter(
            x=[r["Date"] for r in _c], y=[r["NetAfterCosts"] / capital * 100 for r in _c],
            name="2x MTF, after costs", mode="lines", line=dict(width=2)))
        _fig.update_layout(height=380, margin=dict(l=10, r=10, t=30, b=10),
                           yaxis_title="Client's money, indexed to 100",
                           legend=dict(orientation="h", y=1.12))
        st.plotly_chart(_fig, use_container_width=True)

        # -- how much stock per rupee --------------------------------------
        c1, c2, c3 = st.columns(3)
        c1.metric("Stock held per Rs 1 of client money (avg)", f"{_avg_lev:.2f}x")
        c2.metric("Lowest", f"{min(_ratios):.2f}x" if _ratios else "n/a")
        c3.metric("Highest", f"{max(_ratios):.2f}x" if _ratios else "n/a")
        st.caption(
            "'2x MTF' describes how each stock purchase is funded - half own "
            "money, half borrowed. It is not the whole-portfolio exposure. Any "
            "cash or liquid-ETF weight is never borrowed against, so a portfolio "
            "carrying idle cash will show well under 2.00x here. That is expected, "
            "not an error.")

        # -- sensitivity ---------------------------------------------------
        with st.expander("What if the rates were different?"):
            _be = mtf.breakeven_interest(_rows, capital, _cash_ret,
                                         pledge=mtf_pledge,
                                         brokerage_pct=brokerage_pct,
                                         statutory_pct=statutory_pct)
            if _be is None:
                st.markdown(
                    '<div class="warn">At these brokerage and statutory rates, '
                    'borrowing does not beat the cash model at any interest rate '
                    'over this period.</div>', unsafe_allow_html=True)
            else:
                st.metric("Break-even MTF interest rate", f"{_be * 100:.2f}% p.a.",
                          help="Above this rate the client would have been better "
                               "off without MTF, on this strategy over this period.")
            _int_tbl = pd.DataFrame([
                dict(Rate_pct_pa=rt * 100,
                     MTF_return_pct=mtf.roe_after_costs(
                         _rows, capital, interest_pa=rt, pledge=mtf_pledge,
                         brokerage_pct=brokerage_pct,
                         statutory_pct=statutory_pct) * 100,
                     Cash_model_pct=_cash_ret * 100,
                     Advantage_pp=(mtf.roe_after_costs(
                         _rows, capital, interest_pa=rt, pledge=mtf_pledge,
                         brokerage_pct=brokerage_pct,
                         statutory_pct=statutory_pct) - _cash_ret) * 100)
                for rt in (0.10, 0.12, 0.14, 0.15, 0.16, 0.18, 0.20)])
            st.markdown("**If the MTF interest rate changed**")
            st.dataframe(_int_tbl, use_container_width=True, hide_index=True)

            _brok_tbl = []
            for bk in (0.02, 0.05, 0.10, 0.15, 0.20, 0.30):
                _cm = (_cash_gross_end - _cash_turnover
                       * (bk + statutory_pct) / 100.0) / capital - 1
                _mm = mtf.roe_after_costs(_rows, capital,
                                          interest_pa=mtf_rate / 100.0,
                                          pledge=mtf_pledge, brokerage_pct=bk,
                                          statutory_pct=statutory_pct)
                _brok_tbl.append(dict(Brokerage_pct_per_side=bk,
                                      Cash_model_pct=_cm * 100,
                                      MTF_return_pct=_mm * 100,
                                      Advantage_pp=(_mm - _cm) * 100))
            st.markdown("**If the brokerage rate changed**")
            st.dataframe(pd.DataFrame(_brok_tbl), use_container_width=True,
                         hide_index=True)
            st.caption(
                "Both models pay brokerage on every trade, and the MTF book trades "
                "about double the value, so brokerage is usually a bigger lever on "
                "the final answer than the MTF interest rate. Worth checking which "
                "rate is actually negotiable.")

        # -- monthly -------------------------------------------------------
        with st.expander("Month by month"):
            _md = pd.DataFrame([
                dict(Date=r["Date"], Cash=n["NAV"], Mtf=r["NetAfterCosts"])
                for r, n in zip(_c, nav_rows)])
            _md["M"] = pd.to_datetime(_md["Date"]).dt.to_period("M").astype(str)
            _mrows, _pc, _pm = [], capital, capital
            for _m, _g in _md.groupby("M", sort=True):
                _ec, _em = _g["Cash"].iloc[-1], _g["Mtf"].iloc[-1]
                _mrows.append(dict(Month=_m,
                                   Cash_model_pct=(_ec / _pc - 1) * 100,
                                   MTF_pct=(_em / _pm - 1) * 100,
                                   Difference_pp=((_em / _pm) - (_ec / _pc)) * 100))
                _pc, _pm = _ec, _em
            _mdf = pd.DataFrame(_mrows)
            st.dataframe(_mdf, use_container_width=True, hide_index=True)
            _up = _mdf[_mdf.Cash_model_pct > 0]
            _dn = _mdf[_mdf.Cash_model_pct <= 0]
            st.caption(
                f"In the {len(_up)} positive months MTF added "
                f"{_up.Difference_pp.mean():+.2f}pp on average; in the {len(_dn)} "
                f"flat or negative months it cost {_dn.Difference_pp.mean():+.2f}pp. "
                f"Interest is charged every calendar day regardless of whether the "
                f"month worked, so a flat month is a small loss under MTF.")

        # -- accuracy check ------------------------------------------------
        with st.expander("Accuracy check"):
            _ok = abs(_audit["gap"]) < 1.0
            st.dataframe(pd.DataFrame([
                ("Profit locked in on stocks already sold", _audit["realised"]),
                ("Profit on stocks still held", _audit["unrealised"]),
                ("Client money + total profit (what it should be)",
                 capital + _audit["realised"] + _audit["unrealised"]),
                ("What the model reports", _rows[-1]["NetWorth"]),
                ("Difference", _audit["gap"]),
            ], columns=["Item", "Rupees"]), use_container_width=True,
                hide_index=True,
                column_config={"Rupees": st.column_config.NumberColumn(format="%.2f")})
            if _ok:
                st.caption(
                    "Every trade's profit and loss was rebuilt a second, "
                    "independent way (matching sells against the oldest buys "
                    "first) and agrees with the model to the rupee.")
            else:
                st.markdown(
                    '<div class="err">The independent recalculation does not '
                    'agree with the model. Do not use these MTF figures until '
                    'this is resolved.</div>', unsafe_allow_html=True)

        # -- daily detail --------------------------------------------------
        with st.expander("Daily detail"):
            _dd = pd.DataFrame([dict(
                Date=r["Date"],
                CashModelNAV=n["NAV"],
                StockValue=r["GrossAssets"],
                CashBalance=r["Cash"],
                BorrowedAmount=r["TotalLeverage"],
                ClientMoneyValue=r["NetWorth"],
                StockPerRupee=r["StockPerRupee"],
                Positions=r["Positions"],
                CostsToDate=r["AllCosts"],
                ValueAfterCosts=r["NetAfterCosts"],
            ) for r, n in zip(_c, nav_rows)])
            st.dataframe(_dd, use_container_width=True, hide_index=True, height=420)
            st.download_button(
                "Download MTF daily detail (CSV)",
                _dd.to_csv(index=False).encode(),
                file_name=f"{pname.replace(' ', '_')}_MTF_daily.csv",
                mime="text/csv")

# -- corporate-action commit helper -------------------------------------------

typed_now = engine.parse_override_lines(ca_override_txt)[0]
if typed_now:
    with st.expander("Save these corporate-action factors permanently"):
        st.markdown("Streamlit Cloud wipes its filesystem on every redeploy, so the "
                    "app cannot store these itself. Commit the lines below to "
                    "`ca_overrides.csv` at the repo root and every future run, for "
                    "everyone, picks them up automatically.")
        lines = ["symbol,ex_date,value,note"]
        for kk, vv in sorted(typed_now.items()):
            val = (f"{vv['true_ret'] * 100:+.2f}%" if isinstance(vv, dict)
                   else f"{vv}")
            lines.append(engine.override_csv_line(kk[0], kk[1], val, "demerger"))
        st.code("\n".join(lines), language="csv")

# -- export -------------------------------------------------------------------

st.markdown('<div class="sec">Export</div>', unsafe_allow_html=True)

stem = (pname or "portfolio").replace(" ", "_")
ca_applied = [a for a in alerts if "ADJUSTED" in str(a.get("Type", ""))
              and "NOT" not in str(a.get("Type", ""))]

if "nav_excel" not in st.session_state:
    stats_df = pd.DataFrame(
        [{"Section": "Risk", "Metric": k, "Value": v} for k, v in risk.items()]
        + [{"Section": "Cost", "Metric": k, "Value": v} for k, v in costs.items()]
        + [{"Section": "Cash", "Metric": k, "Value": v} for k, v in cashd.items()])
    cover = [
        ("Portfolio", pname),
        ("Period", f"{first['Date']:%d %b %Y} to {last['Date']:%d %b %Y}"),
        ("Sessions", len(nav_rows)),
        ("Starting capital", capital),
        ("Final NAV", round(last["NAV"], 2)),
        ("Return excl. dividends %", round(total_return, 2)),
        ("Return incl. dividends %", round(total_return_tr, 2)),
        ("Dividends received", round(div_total, 2)),
        ("Benchmark", bname),
        ("Outperformance pp", round(outperf, 2) if outperf is not None else "n/a"),
        ("Rebalance tolerance pp", gate_pp),
        ("Rebalance all holdings", "Yes" if force_retarget else "No"),
        ("Corporate actions", "Adjusted" if ca_on else "Not adjusted"),
        ("CA basis", ca_mode),
        ("Demerger policy", dm_policy),
        ("CA adjustments applied", len(ca_applied)),
        ("Brokerage % per side", brokerage_pct),
        ("Statutory % per side", statutory_pct),
        ("Alerts raised", len(alerts_df)),
        ("Price source", "NSE bhavcopy (official closing prices)"),
        ("Basis", "Price return"),
        ("Generated", date.today().strftime("%d %b %Y")),
    ]
    st.session_state["nav_excel"] = to_excel([
        ("Performance", perf_df), ("Performance incl Div", perf_tr_df),
        ("Risk and Costs", stats_df), ("Contributors", attrib_df),
        ("Dividends", div_df), ("Daily NAV", nav_df),
        ("Final Holdings", holdings_export), ("Holdings Wt Matrix", wt_matrix),
        ("Holdings Qty Matrix", qty_matrix), ("Trades", trades_df),
        ("Alerts", alerts_df), ("Reconciliation", recon_df),
    ], cover)
    st.session_state["nav_csv"] = nav_df.to_csv(index=False).encode()

if PDF_OK and "nav_pdf" not in st.session_state:
    try:
        st.session_state["nav_pdf"] = pdf_report.build_factsheet(
            nav_rows=nav_rows, perf_df=perf_df, holdings_df=holdings_export,
            risk=risk, costs=costs, cashd=cashd, attrib_df=attrib_df,
            bench=bench, bench_name=bname, capital=capital,
            portfolio_name=pname, palette=P, div_total=div_total, external=False)
    except Exception as e:
        st.session_state["nav_pdf"] = None
        st.caption(f"Factsheet could not be generated: {e}")

e1, e2, e3 = st.columns(3, gap="small")
with e1:
    st.download_button("Download factsheet (PDF)",
                       data=st.session_state.get("nav_pdf") or b"",
                       file_name=f"{stem}_factsheet_{to_date:%d%b%Y}.pdf",
                       mime="application/pdf", use_container_width=True,
                       type="primary",
                       disabled=not st.session_state.get("nav_pdf"))
with e2:
    st.download_button("Download workbook (Excel)",
                       data=st.session_state["nav_excel"],
                       file_name=f"NAV_{stem}_{to_date:%d%b%Y}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet",
                       use_container_width=True)
with e3:
    st.download_button("Download NAV series (CSV)",
                       data=st.session_state["nav_csv"],
                       file_name=f"NAV_{stem}_{to_date:%d%b%Y}.csv",
                       mime="text/csv", use_container_width=True)

# -- disclosure ---------------------------------------------------------------

st.markdown(
    '<div class="foot"><strong>Basis and sources.</strong> Simulated performance of '
    'the model portfolio for a client investing at inception, driven by the research '
    'advice log. Prices are official NSE closing prices from the exchange bhavcopy '
    'archive; trades execute at the prices recorded in the advice log, with the '
    'session close substituted where a log price is missing, zero, or predates the '
    'start date. Returns are price returns; the dividend-inclusive basis credits cash '
    'on the ex-date and is shown separately. The benchmark is a price index and '
    'excludes its constituents&#39; dividends. Corporate actions are read from NSE&#39;s '
    'official calendar: splits and bonuses use the exact published terms, demergers '
    'require a factor supplied by the user. Cash earns no return. Quantities are '
    'floored to whole shares, so smaller ticket sizes carry a larger rounding drag. '
    'This is a simulation and not the record of any individual client account; actual '
    'results vary with entry date, execution and costs. Past performance is not '
    'indicative of future results. Internal use only.</div>',
    unsafe_allow_html=True)
