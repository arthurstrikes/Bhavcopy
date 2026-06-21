import streamlit as st
import pandas as pd
import yfinance as yf
from io import BytesIO
from datetime import datetime, timedelta, date
import time

# ── PAGE CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="Bhavcopy – NSE Close Prices",
    page_icon="📋",
    layout="wide"
)

st.markdown("""
<style>
    .title { font-size: 2rem; font-weight: 700; color: #1F3864; }
    .subtitle { font-size: 1rem; color: #555; margin-bottom: 1.5rem; }
    .info-box { background: #EBF5FB; border-left: 4px solid #2E86C1;
                padding: 0.8rem 1rem; border-radius: 4px; font-size: 0.9rem; margin-bottom: 1rem; }
    .success-box { background: #EAFAF1; border-left: 4px solid #27AE60;
                   padding: 0.8rem 1rem; border-radius: 4px; font-size: 0.9rem; }
    .warn-box { background: #FEF9E7; border-left: 4px solid #F39C12;
                padding: 0.8rem 1rem; border-radius: 4px; font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="title">📋 Bhavcopy — NSE/BSE Closing Prices</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Fetch NSE/BSE stock and index closing prices. Paste from Excel or type symbols directly.</div>', unsafe_allow_html=True)

# ── HOW TO USE ───────────────────────────────────────────────
with st.expander("📖 How to use", expanded=False):
    st.markdown("""
**Two ways to get prices:**

---

**📋 Matrix Mode** — Best when you already have a table in Excel

Set up your table with stock/index names in the first column and dates in the first row:
```
Symbol      29-Sep-25   02-Nov-25   17-Apr-26
RELIANCE
INFY
NIFTY50
SENSEX
```
Select all → Copy → Paste into the Matrix tab → click **Fetch Closing Prices**.

💡 **No dates in your table?** Just paste the symbol list and the app will automatically fetch the last available trading day's closes.

---

**⚡ Quick Fetch** — Best for quick lookups without needing Excel

Type stock or index names (one per line or comma-separated), pick your date option, and fetch.
Options: **Last trading day** (default), **specific dates**, or a **date range**.

---

**Options (work in both modes):**

| Option | What it does |
|---|---|
| **Fill holidays** | Holiday/weekend cells show nearest prior trading day's close, marked `*` in Excel. Off = leave blank. |
| **Adjusted prices** | Toggle between unadjusted (default) and split/dividend-adjusted closes. |

---

**Supported index names:**

| Category | Type exactly as shown |
|---|---|
| NSE Broad | `NIFTY50`, `NIFTY100`, `NIFTY200`, `NIFTY500`, `NIFTYNEXT50` |
| NSE Mid/Small | `NIFTYMIDCAP100`, `NIFTYMID100`, `NIFTYSMALLCAP100`, `NIFTYSC100` |
| NSE Sectoral | `BANKNIFTY`, `NIFTYIT`, `NIFTYAUTO`, `NIFTYPHARMA`, `NIFTYFMCG`, `NIFTYMETAL`, `NIFTYREALTY`, `NIFTYENERGY`, `NIFTYPSUBANK`, `FINNIFTY` |
| NSE Composite† | `NIFTYMIDSML400`, `NIFTY200MOM30`, `NIFTYLARGEMID250` |
| BSE Broad | `SENSEX`, `BSE100`, `BSE200`, `BSE500` |
| BSE Mid/Small | `BSEMIDCAP`, `BSESMALLCAP` |
| BSE Sectoral | `BSEBANK`, `BSEIT`, `BSEAUTO`, `BSEPHARMA`, `BSEFMCG`, `BSEMETAL`, `BSEREALTY`, `BSEENERGY` |
| Volatility | `INDIAVIX`, `VIX` |

† NSE Composite indices are not available on Yahoo Finance. The app will flag them and direct you to download from NSE India manually.

**Notes:**
- Use exact NSE ticker symbols for stocks: `BAJAJ-AUTO`, `M&M`, `L&T` etc.
- Large batches (50+ stocks × many dates) take 2–4 minutes.
- Adjusted prices affect historical values — use unadjusted for most report work.
    """)

# ── UTILITY FUNCTIONS ────────────────────────────────────────
def last_trading_day():
    """Most recent completed weekday (yesterday or earlier). Does not account for market holidays."""
    d = date.today() - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def parse_date_flexible(date_str):
    date_str = str(date_str).strip()
    if not date_str or date_str.lower() in ('nan', 'none', ''):
        return None
    formats = [
        "%d-%b-%y", "%d-%b-%Y",
        "%d/%m/%Y", "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y", "%d-%m-%y",
        "%b %d %Y", "%b %d, %Y",
        "%d %b %Y", "%d %b %y",
        "%m/%d/%Y",
        "%Y%m%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            pass
    try:
        return pd.to_datetime(date_str, dayfirst=True).date()
    except:
        return None

# ── PARSE MATRIX INPUT ───────────────────────────────────────
def parse_input(raw):
    """
    Returns: (symbols, dates_with_labels, unparsed, auto_date)
    On error: (None, error_string, [], False)
    auto_date=True means no dates were found — last trading day was auto-applied.
    """
    lines = raw.strip().split('\n')
    if len(lines) < 1:
        return None, "Nothing pasted.", [], False

    header     = lines[0].split('\t')
    first_cell = header[0].strip()

    if first_cell and parse_date_flexible(first_cell):
        date_strings = [h.strip() for h in header if h.strip()]
        symbol_lines = lines
    else:
        date_strings = [h.strip() for h in header[1:] if h.strip()]
        symbol_lines = lines[1:]

    dates, unparsed = [], []
    for ds in date_strings:
        d = parse_date_flexible(ds)
        if d:
            dates.append((ds, d))
        else:
            unparsed.append(ds)

    # Zero dates → auto-use last trading day
    auto_date = False
    if not dates:
        ltd = last_trading_day()
        dates = [(ltd.strftime('%d-%b-%Y'), ltd)]
        auto_date = True

    symbols = []
    for line in symbol_lines:
        parts = line.split('\t')
        sym = parts[0].strip().upper()
        if sym:
            symbols.append(sym)

    if not symbols:
        return None, "No stock symbols found in the first column.", [], False

    return symbols, dates, unparsed, auto_date

# ── PARSE QUICK FETCH SYMBOLS ────────────────────────────────
def parse_quick_symbols(raw):
    """Accepts newline or comma-separated symbol list."""
    raw = raw.replace(',', '\n')
    return [s.strip().upper() for s in raw.split('\n') if s.strip()]

# ── TICKER ALIAS MAP ─────────────────────────────────────────
TICKER_ALIASES = {
    "GMDC":         "GMDCLTD",
    "LT":           "LT",
    "M&M":          "M&M",
    "MM":           "M&M",
    "BAJAJ AUTO":   "BAJAJ-AUTO",
    "BAJAJ_AUTO":   "BAJAJ-AUTO",
    "NMDCSTEEL":    "NSLNISP",
    "MOTHERSUMI":   "MOTHERSON",
    "LTIM":         "LTIM",
    "LTIMINDTREE":  "LTIM",
    "HINDZINC":     "HINDZINC",
}

# ── INDEX TICKER MAP ─────────────────────────────────────────
INDEX_TICKERS = {
    # ── NSE Broad Market ─────────────────────────────────────
    "NIFTY50":          "^NSEI",
    "NIFTY 50":         "^NSEI",
    "NIFTY":            "^NSEI",
    "NIFTY100":         "^CNX100",
    "NIFTY 100":        "^CNX100",
    "NIFTY200":         "^CNX200",
    "NIFTY 200":        "^CNX200",
    "NIFTY500":         "^CRSLDX",
    "NIFTY 500":        "^CRSLDX",
    "NIFTYNEXT50":      "^NSMIDCP50",
    "NIFTY NEXT 50":    "^NSMIDCP50",
    "NIFTYJR":          "^NSMIDCP50",

    # ── NSE Midcap / Smallcap ────────────────────────────────
    "NIFTYMIDCAP50":    "^NSEMDCP50",
    "NIFTYMID50":       "^NSEMDCP50",
    "NIFTYMIDCAP100":   "^CNXMIDCAP",
    "NIFTYMID100":      "^CNXMIDCAP",
    "NIFTYMIDCAP150":   "NIFTYMIDCAP150.NS",
    "NIFTYMID150":      "NIFTYMIDCAP150.NS",
    "NIFTYSMALLCAP50":  "^CNXSC",
    "NIFTYSC50":        "^CNXSC",
    "NIFTYSMALLCAP100": "^CNXSC",
    "NIFTYSC100":       "^CNXSC",
    "NIFTYSMALLCAP250": "NIFTYSMALLCAP250.NS",
    "NIFTYSC250":       "NIFTYSMALLCAP250.NS",
    "NIFTYMICROCAP250": "NIFTYMICROCAP250.NS",

    # ── NSE Sectoral ─────────────────────────────────────────
    "BANKNIFTY":        "^NSEBANK",
    "NIFTYBANK":        "^NSEBANK",
    "NIFTY BANK":       "^NSEBANK",
    "NIFTYIT":          "^CNXIT",
    "NIFTY IT":         "^CNXIT",
    "NIFTYAUTO":        "^CNXAUTO",
    "NIFTY AUTO":       "^CNXAUTO",
    "NIFTYPHARMA":      "^CNXPHARMA",
    "NIFTY PHARMA":     "^CNXPHARMA",
    "NIFTYFMCG":        "^CNXFMCG",
    "NIFTY FMCG":       "^CNXFMCG",
    "NIFTYMETAL":       "^CNXMETAL",
    "NIFTY METAL":      "^CNXMETAL",
    "NIFTYREALTY":      "^CNXREALTY",
    "NIFTY REALTY":     "^CNXREALTY",
    "NIFTYENERGY":      "^CNXENERGY",
    "NIFTY ENERGY":     "^CNXENERGY",
    "NIFTYINFRA":       "^CNXINFRA",
    "NIFTY INFRA":      "^CNXINFRA",
    "NIFTYMEDIA":       "^CNXMEDIA",
    "NIFTY MEDIA":      "^CNXMEDIA",
    "NIFTYPSUBANK":     "^CNXPSUBANK",
    "NIFTY PSU BANK":   "^CNXPSUBANK",
    "NIFTYPSU":         "^CNXPSE",
    "NIFTY PSU":        "^CNXPSE",
    "NIFTYPSE":         "^CNXPSE",
    "NIFTYFINSERVICE":  "NIFTYFINSERVICE.NS",
    "NIFTY FIN SERVICE":"NIFTYFINSERVICE.NS",
    "FINNIFTY":         "NIFTYFINSERVICE.NS",
    "NIFTYHEALTHCARE":  "^CNXPHARMA",
    "NIFTYCONSUMER":    "NIFTYCONSUMPTION.NS",
    "NIFTYOILGAS":      "NIFTYOILGAS.NS",
    "NIFTYMFG":         "NIFTYMFG.NS",
    "NIFTYDEFENCE":     "NIFTYDEFENCE.NS",

    # ── NSE Strategy / Other ─────────────────────────────────
    "NIFTYALPHA50":          "NIFTYALPHA50.NS",
    "NIFTYDIVIDEND":         "^CNXDIVID",
    "NIFTYGROWTH":           "^CNXGRW25",
    "NIFTYCPSE":             "^CNXCPSE",
    "INDIA VIX":             "^INDIAVIX",
    "INDIAVIX":              "^INDIAVIX",
    "VIX":                   "^INDIAVIX",

    # ── NSE Composite / Factor (chart-only on YF; in YF_UNSUPPORTED_INDICES) ─
    "NIFTYMIDSML400":        "NIFTYMIDSML400.NS",
    "NIFTY MIDSML 400":      "NIFTYMIDSML400.NS",
    "NIFTYMIDSMALL400":      "NIFTYMIDSML400.NS",
    "NIFTY MID SMALL 400":   "NIFTYMIDSML400.NS",
    "NIFTY200MOM30":         "NIFTY200MOMENTM30.NS",
    "NIFTY200MOMENTUM30":    "NIFTY200MOMENTM30.NS",
    "NIFTY200 MOMENTUM 30":  "NIFTY200MOMENTM30.NS",
    "NIFTYLARGEMID250":      "NIFTY_LARGEMID250.NS",
    "NIFTYLARGEMIDCAP250":   "NIFTY_LARGEMID250.NS",
    "NIFTY LARGEMIDCAP 250": "NIFTY_LARGEMID250.NS",
    "NIFTYLARGMID250":       "NIFTY_LARGEMID250.NS",

    # ── BSE Broad Market ─────────────────────────────────────
    "SENSEX":           "^BSESN",
    "BSE SENSEX":       "^BSESN",
    "BSE30":            "^BSESN",
    "BSE100":           "BSE-100.BO",
    "BSE 100":          "BSE-100.BO",
    "BSE200":           "BSE-200.BO",
    "BSE 200":          "BSE-200.BO",
    "BSE500":           "BSE-500.BO",
    "BSE 500":          "BSE-500.BO",

    # ── BSE Midcap / Smallcap ────────────────────────────────
    "BSEMIDCAP":        "BSE-MIDCAP.BO",
    "BSE MIDCAP":       "BSE-MIDCAP.BO",
    "BSESMALLCAP":      "BSE-SMLCAP.BO",
    "BSE SMALLCAP":     "BSE-SMLCAP.BO",
    "BSE SMLCAP":       "BSE-SMLCAP.BO",
    "BSEMICROCAP":      "BSE-MICROCAP.BO",
    "BSE MICROCAP":     "BSE-MICROCAP.BO",
    "BSELARGECAP":      "BSE-LARGECAP.BO",
    "BSE LARGECAP":     "BSE-LARGECAP.BO",

    # ── BSE Sectoral ─────────────────────────────────────────
    "BSEBANK":          "BSE-BANK.BO",
    "BSE BANK":         "BSE-BANK.BO",
    "BSEIT":            "BSE-IT.BO",
    "BSE IT":           "BSE-IT.BO",
    "BSEAUTO":          "BSE-AUTO.BO",
    "BSE AUTO":         "BSE-AUTO.BO",
    "BSEPHARMA":        "BSE-HEALTHCARE.BO",
    "BSE PHARMA":       "BSE-HEALTHCARE.BO",
    "BSEHEALTHCARE":    "BSE-HEALTHCARE.BO",
    "BSE HEALTHCARE":   "BSE-HEALTHCARE.BO",
    "BSEFMCG":          "BSE-FMCG.BO",
    "BSE FMCG":         "BSE-FMCG.BO",
    "BSEMETAL":         "BSE-METAL.BO",
    "BSE METAL":        "BSE-METAL.BO",
    "BSEREALTY":        "BSE-REALTY.BO",
    "BSE REALTY":       "BSE-REALTY.BO",
    "BSEENERGY":        "BSE-ENERGY.BO",
    "BSE ENERGY":       "BSE-ENERGY.BO",
    "BSEOILGAS":        "BSE-OIL&GAS.BO",
    "BSE OIL GAS":      "BSE-OIL&GAS.BO",
    "BSECAPGOODS":      "BSE-CARGDS.BO",
    "BSE CAP GOODS":    "BSE-CARGDS.BO",
    "BSEPSU":           "BSE-PSU.BO",
    "BSE PSU":          "BSE-PSU.BO",
    "BSEFINANCE":       "BSE-FIN.BO",
    "BSE FINANCE":      "BSE-FIN.BO",
    "BSEPOWER":         "BSE-POWER.BO",
    "BSE POWER":        "BSE-POWER.BO",
    "BSETECK":          "BSE-TECK.BO",
    "BSE TECK":         "BSE-TECK.BO",
    "BSECONSUMERDURABLES": "BSE-CONSDUR.BO",
    "BSE CONSUMER DURABLES": "BSE-CONSDUR.BO",
    "BSEINDUSTRIALS":   "BSE-INDUS.BO",
    "BSE INDUSTRIALS":  "BSE-INDUS.BO",
    "BSETELECOMMUNICATION": "BSE-TELECOM.BO",
    "BSE TELECOM":      "BSE-TELECOM.BO",
    "BSEUTILS":         "BSE-UTILS.BO",
    "BSE UTILITIES":    "BSE-UTILS.BO",
}

# ── YF-UNSUPPORTED INDICES ───────────────────────────────────
# Yahoo Finance shows charts but serves no downloadable historical OHLCV for these.
# Listed in INDEX_TICKERS so is_index() recognises them and no double .NS is appended.
# fetch_prices() detects them here and shows an actionable message.
YF_UNSUPPORTED_INDICES = {
    "NIFTYMIDSML400", "NIFTY MIDSML 400", "NIFTYMIDSMALL400", "NIFTY MID SMALL 400",
    "NIFTY200MOM30", "NIFTY200MOMENTUM30", "NIFTY200 MOMENTUM 30",
    "NIFTYLARGEMID250", "NIFTYLARGEMIDCAP250", "NIFTY LARGEMIDCAP 250", "NIFTYLARGMID250",
}

def is_index(symbol):
    return symbol.strip().upper() in INDEX_TICKERS

def to_yf_ticker(symbol):
    symbol = symbol.strip().upper()
    if symbol in INDEX_TICKERS:
        return INDEX_TICKERS[symbol]
    symbol = TICKER_ALIASES.get(symbol, symbol)
    # Guard: don't double-suffix if analyst typed a full Yahoo ticker already
    if symbol.endswith('.NS') or symbol.endswith('.BO') or symbol.startswith('^'):
        return symbol
    return symbol + ".NS"

# ── FETCH LOGIC ──────────────────────────────────────────────
MAX_RETRIES   = 3
RETRY_DELAY   = 4
REQUEST_DELAY = 0.5

def fetch_single(ticker, fetch_start, fetch_end, adjusted=False):
    """Download data for one ticker with retries. Returns DataFrame or raises."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            df = yf.download(
                ticker,
                start=fetch_start,
                end=fetch_end,
                interval="1d",
                progress=False,
                auto_adjust=adjusted,
                actions=False,
            )
            if not df.empty:
                return df
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
            else:
                raise e
    return pd.DataFrame()


def fetch_prices(symbols, date_objects, fill_holidays=False, adjusted=False):
    """
    Returns: results, failed, failed_errors, holiday_fills
    holiday_fills: set of date objects where prior-day fill was applied.
    """
    min_date = min(date_objects)
    max_date = max(date_objects)

    fetch_start = (min_date - timedelta(days=10)).strftime("%Y-%m-%d")
    fetch_end   = (max_date + timedelta(days=12)).strftime("%Y-%m-%d")

    results       = {}
    failed        = []
    failed_errors = {}
    holiday_fills = set()   # dates where prior-day fill was applied

    progress = st.progress(0, text="Starting…")
    total = len(symbols)

    for i, symbol in enumerate(symbols):
        progress.progress(int(i / total * 100), text=f"Fetching {symbol}  ({i+1} of {total})")
        ticker = to_yf_ticker(symbol)

        try:
            df = fetch_single(ticker, fetch_start, fetch_end, adjusted=adjusted)

            if df.empty:
                failed.append(symbol)
                if symbol.upper() in YF_UNSUPPORTED_INDICES:
                    failed_errors[symbol] = (
                        "Not available via Yahoo Finance — download historical data manually from "
                        "NSE India (nseindia.com → Reports → Index Historical Data) "
                        "and paste closes into your Excel"
                    )
                elif is_index(symbol):
                    failed_errors[symbol] = f"No data for index {ticker} — may not be available on Yahoo Finance for this date range"
                else:
                    failed_errors[symbol] = f"No data returned for {ticker} — verify NSE ticker"
                results[symbol] = {}
                time.sleep(REQUEST_DELAY)
                continue

            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)

            df.index = pd.to_datetime(df.index).date
            close_map = {}

            for dt in date_objects:
                if dt in df.index:
                    val = df.loc[dt, "Close"]
                    close_map[dt] = round(float(val), 2) if pd.notna(val) else None
                elif fill_holidays:
                    # Walk back up to 7 days to find nearest prior trading day
                    filled = False
                    for days_back in range(1, 8):
                        prior = dt - timedelta(days=days_back)
                        if prior in df.index:
                            val = df.loc[prior, "Close"]
                            close_map[dt] = round(float(val), 2) if pd.notna(val) else None
                            holiday_fills.add(dt)
                            filled = True
                            break
                    if not filled:
                        close_map[dt] = None
                else:
                    close_map[dt] = None   # holiday/weekend → blank

            results[symbol] = close_map

        except Exception as e:
            failed.append(symbol)
            failed_errors[symbol] = str(e)[:120]
            results[symbol] = {}

        time.sleep(REQUEST_DELAY)

    progress.progress(100, text="Done!")
    return results, failed, failed_errors, holiday_fills


# ── BUILD OUTPUT ─────────────────────────────────────────────
def build_output(symbols, dates_with_labels, price_data):
    rows = []
    for sym in symbols:
        row = {"Symbol": sym}
        for orig, dt in dates_with_labels:
            row[orig] = price_data.get(sym, {}).get(dt, None)
        rows.append(row)
    return pd.DataFrame(rows)


# ── EXCEL EXPORT ─────────────────────────────────────────────
def to_excel(df, holiday_fills=None, dates_with_labels=None):
    """
    holiday_fills: set of date objects where prior-day fill was applied.
    dates_with_labels: list of (label_str, date_obj) to map fills to column names.
    Holiday columns get amber headers and a * suffix in the label.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Rename holiday columns: append * to header label
    holiday_label_set = set()
    if holiday_fills and dates_with_labels:
        for orig_label, dt in dates_with_labels:
            if dt in holiday_fills:
                holiday_label_set.add(orig_label)
        if holiday_label_set:
            df = df.rename(columns={lbl: lbl + "*" for lbl in holiday_label_set})

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Bhavcopy")

        ws   = writer.sheets["Bhavcopy"]
        navy  = PatternFill("solid", fgColor="1F3864")
        amber = PatternFill("solid", fgColor="FFF0A0")   # light amber for holiday cols
        thin  = Side(style="thin", color="D0D0D0")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=col_idx)
            is_holiday_col = str(col_name).endswith("*")
            c.fill      = amber if is_holiday_col else navy
            c.font      = Font(name="Arial",
                               color="000000" if is_holiday_col else "FFFFFF",
                               bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = bdr

        # Data rows
        for row in range(2, len(df) + 2):
            c = ws.cell(row=row, column=1)
            c.font   = Font(name="Arial", bold=True, size=10)
            c.border = bdr
            for col in range(2, len(df.columns) + 1):
                c = ws.cell(row=row, column=col)
                c.font          = Font(name="Arial", size=10)
                c.number_format = "#,##0.00"
                c.alignment     = Alignment(horizontal="right")
                c.border        = bdr

        # Column widths
        ws.column_dimensions["A"].width = 16
        for col in range(2, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 13

        ws.freeze_panes = "B2"

        # Footer note if any holiday fills exist
        if holiday_label_set:
            note_row = len(df) + 3
            note_cell = ws.cell(row=note_row, column=1)
            note_cell.value = "* Holiday or weekend — nearest prior trading day close shown"
            note_cell.font  = Font(name="Arial", italic=True, size=9, color="888888")

    buf.seek(0)
    return buf


# ── SHARED FETCH + DISPLAY ───────────────────────────────────
def run_fetch_and_display(symbols, dates_with_labels, fill_holidays, adjusted, auto_date=False):
    """Execute fetch and render results. Called from both Matrix and Quick Fetch tabs."""
    date_objects  = [d for _, d in dates_with_labels]
    index_syms    = [s for s in symbols if is_index(s)]
    stock_syms    = [s for s in symbols if not is_index(s)]
    breakdown     = ""
    if index_syms:
        breakdown = f"&nbsp;|&nbsp; <strong>{len(stock_syms)}</strong> stocks + <strong>{len(index_syms)}</strong> indexes"

    auto_note = ""
    if auto_date:
        auto_note = "&nbsp;|&nbsp; ⚡ No dates found — auto-fetching <strong>last trading day</strong>"

    st.markdown(f"""
    <div class="success-box">
    ✅ Parsed — <strong>{len(symbols)} rows</strong> × <strong>{len(dates_with_labels)} dates</strong>
    {breakdown}
    &nbsp;|&nbsp; Range: <strong>{min(date_objects).strftime('%d-%b-%Y')}</strong>
    to <strong>{max(date_objects).strftime('%d-%b-%Y')}</strong>
    {auto_note}
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Symbols detected:**")
        st.write(", ".join(symbols))
    with c2:
        st.markdown("**Dates detected:**")
        preview_dates = [d.strftime('%d-%b-%Y') for _, d in dates_with_labels[:12]]
        suffix = f" … +{len(dates_with_labels)-12} more" if len(dates_with_labels) > 12 else ""
        st.write(", ".join(preview_dates) + suffix)

    st.markdown("---")

    price_data, failed, failed_errors, holiday_fills = fetch_prices(
        symbols, date_objects, fill_holidays=fill_holidays, adjusted=adjusted
    )
    output_df = build_output(symbols, dates_with_labels, price_data)

    # Persist in session_state — survives Streamlit reruns (e.g. on download click)
    st.session_state["output_df"]       = output_df
    st.session_state["output_symbols"]  = symbols
    st.session_state["output_dates"]    = dates_with_labels
    st.session_state["holiday_fills"]   = holiday_fills

    total_cells  = len(symbols) * len(dates_with_labels)
    filled_cells = output_df.iloc[:, 1:].notna().sum().sum()
    holiday_note = f"&nbsp;|&nbsp; <strong>{len(holiday_fills)}</strong> holiday date(s) filled with prior close" if holiday_fills else ""

    st.markdown(f"""
    <div class="success-box">
    ✅ <strong>Complete</strong> &nbsp;|&nbsp;
    <strong>{filled_cells:,}</strong> prices fetched &nbsp;|&nbsp;
    <strong>{total_cells - filled_cells:,}</strong> blanks (holidays/weekends) &nbsp;|&nbsp;
    <strong>{len(failed)}</strong> symbols failed
    {holiday_note}
    </div>
    """, unsafe_allow_html=True)

    if failed:
        failed_lines = "<br>".join(
            f"<strong>{s}</strong>: {failed_errors.get(s, 'unknown error')}" for s in failed
        )
        st.markdown(f'<div class="warn-box">⚠️ Failed symbols:<br>{failed_lines}</div>', unsafe_allow_html=True)

    # Preview — cap at 15 columns for display only
    st.markdown("### Preview")
    preview_cols = ["Symbol"] + [orig for orig, _ in dates_with_labels[:15]]
    # Use renamed df if holiday fills exist (so * shows in preview too)
    preview_df = output_df.copy()
    if holiday_fills and dates_with_labels:
        preview_df = preview_df.rename(columns={
            orig: orig + "*" for orig, dt in dates_with_labels if dt in holiday_fills
        })
        preview_cols = [c + "*" if c in {orig for orig, dt in dates_with_labels if dt in holiday_fills} else c
                        for c in preview_cols]

    st.dataframe(
        preview_df[[c for c in preview_cols if c in preview_df.columns]].style.format(
            {col: "{:,.2f}" for col in preview_cols[1:] if col in preview_df.columns},
            na_rep="—"
        ),
        use_container_width=True,
        height=min(400, (len(symbols) + 1) * 38),
    )
    if len(dates_with_labels) > 15:
        st.caption(f"Showing 15 of {len(dates_with_labels)} date columns. All {len(dates_with_labels)} columns included in the Excel download.")

    # Download — read from session_state to survive reruns
    dl_df    = st.session_state.get("output_df", output_df)
    dl_dates = st.session_state.get("output_dates", dates_with_labels)
    dl_syms  = st.session_state.get("output_symbols", symbols)
    dl_hf    = st.session_state.get("holiday_fills", holiday_fills)

    price_label = "Adjusted" if adjusted else "Unadjusted"
    st.download_button(
        label=f"⬇️ Download — {len(dl_syms)} stocks × {len(dl_dates)} dates ({price_label}) Excel",
        data=to_excel(dl_df, holiday_fills=dl_hf, dates_with_labels=dl_dates),
        file_name=f"Bhavcopy_{datetime.today().strftime('%d%b%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


# ── GLOBAL OPTIONS ───────────────────────────────────────────
st.markdown("### Options")
opt_col1, opt_col2 = st.columns(2)
with opt_col1:
    fill_holidays = st.checkbox(
        "Fill holidays with prior trading day close",
        value=False,
        help="When on, holiday/weekend cells show the nearest prior trading day's close, marked * in Excel. When off, they show blank."
    )
with opt_col2:
    price_type = st.radio(
        "Price type",
        ["Unadjusted (default)", "Adjusted (split/dividend)"],
        horizontal=True,
        help="Unadjusted = raw exchange close. Adjusted = corrected for splits and dividends."
    )
    adjusted = price_type.startswith("Adjusted")

st.markdown("---")

# ── INPUT TABS ───────────────────────────────────────────────
tab_matrix, tab_quick = st.tabs(["📋 Matrix Mode — Paste from Excel", "⚡ Quick Fetch — Type symbols"])

# ════════════════════════════════════════════════════════════
# TAB 1 — MATRIX MODE
# ════════════════════════════════════════════════════════════
with tab_matrix:
    st.markdown('<div class="info-box">Copy your Symbol × Date table from Excel and paste below. First column = stock/index names, first row = dates. Leave price cells empty.<br>💡 <strong>No dates?</strong> Just paste symbol names — the app will fetch last trading day automatically.</div>', unsafe_allow_html=True)

    raw_input = st.text_area(
        label="Paste here",
        height=220,
        placeholder="Symbol\t29-Sep-25\t02-Nov-25\t17-Apr-26\nRELIANCE\t\t\t\nINFY\t\t\t\nNIFTY50\t\t\t",
        label_visibility="collapsed",
        key="raw_input_area"
    )
    # Persist full input across reruns to prevent browser truncation
    if raw_input:
        st.session_state["last_raw_input"] = raw_input
    raw_input = st.session_state.get("last_raw_input", raw_input)

    fetch_matrix = st.button("🔄 Fetch Closing Prices", type="primary",
                              use_container_width=True, key="btn_matrix")

    if fetch_matrix:
        if not raw_input.strip():
            st.error("⚠️ Paste your table above before fetching.")
        else:
            result = parse_input(raw_input)
            if result[0] is None:
                st.error(result[1])
            else:
                symbols, dates_with_labels, unparsed, auto_date = result
                if unparsed:
                    st.markdown(f'<div class="warn-box">⚠️ Skipped unrecognised date values: {", ".join(unparsed)}</div>', unsafe_allow_html=True)
                run_fetch_and_display(symbols, dates_with_labels, fill_holidays, adjusted, auto_date)

    elif raw_input.strip():
        # Live parse preview before fetch
        result = parse_input(raw_input)
        if result[0] is None:
            st.error(result[1])
        else:
            symbols, dates_with_labels, unparsed, auto_date = result
            date_objects = [d for _, d in dates_with_labels]
            index_syms   = [s for s in symbols if is_index(s)]
            stock_syms   = [s for s in symbols if not is_index(s)]
            breakdown    = f"&nbsp;|&nbsp; <strong>{len(stock_syms)}</strong> stocks + <strong>{len(index_syms)}</strong> indexes" if index_syms else ""
            auto_note    = "&nbsp;|&nbsp; ⚡ No dates found — will fetch <strong>last trading day</strong>" if auto_date else ""

            st.markdown(f"""
            <div class="success-box">
            ✅ Parsed — <strong>{len(symbols)} rows</strong> × <strong>{len(dates_with_labels)} dates</strong>
            {breakdown}
            &nbsp;|&nbsp; Range: <strong>{min(date_objects).strftime('%d-%b-%Y')}</strong>
            to <strong>{max(date_objects).strftime('%d-%b-%Y')}</strong>
            {auto_note}
            &nbsp;|&nbsp; Ready — click <strong>Fetch Closing Prices</strong> above.
            </div>
            """, unsafe_allow_html=True)
            if unparsed:
                st.markdown(f'<div class="warn-box">⚠️ Skipped unrecognised date values: {", ".join(unparsed)}</div>', unsafe_allow_html=True)

    with st.expander("📊 See example input format", expanded=False):
        st.dataframe(
            pd.DataFrame({
                "Symbol":    ["RELIANCE","INFY","NIFTY50","SENSEX","BANKNIFTY"],
                "29-Sep-25": [""]*5,
                "02-Nov-25": [""]*5,
                "17-Apr-26": [""]*5,
            }),
            use_container_width=True,
            hide_index=True,
        )
        st.caption("Mix stocks and index names freely. Copy this table structure from Excel.")


# ════════════════════════════════════════════════════════════
# TAB 2 — QUICK FETCH
# ════════════════════════════════════════════════════════════
with tab_quick:
    st.markdown('<div class="info-box">Type stock/index names below. Choose your date option. No Excel needed.</div>', unsafe_allow_html=True)

    qf_symbols_raw = st.text_area(
        "Symbols (one per line or comma-separated)",
        height=160,
        placeholder="RELIANCE\nINFY\nHCLTECH\nNIFTY50\nSENSEX",
        key="qf_symbols"
    )

    qf_date_mode = st.radio(
        "Date",
        ["Last trading day", "Specific dates", "Date range"],
        horizontal=True,
        key="qf_date_mode"
    )

    qf_dates_with_labels = []
    qf_date_error        = None

    if qf_date_mode == "Last trading day":
        ltd = last_trading_day()
        qf_dates_with_labels = [(ltd.strftime('%d-%b-%Y'), ltd)]
        st.caption(f"Will fetch: **{ltd.strftime('%d-%b-%Y')}**")

    elif qf_date_mode == "Specific dates":
        qf_dates_raw = st.text_input(
            "Dates (comma-separated, any format)",
            placeholder="29-Sep-25, 02-Nov-25, 17-Apr-26",
            key="qf_dates_specific"
        )
        if qf_dates_raw.strip():
            for part in qf_dates_raw.replace('\n', ',').split(','):
                part = part.strip()
                if part:
                    d = parse_date_flexible(part)
                    if d:
                        qf_dates_with_labels.append((part, d))
                    else:
                        st.warning(f"Could not parse date: `{part}` — skipped")
        if not qf_dates_with_labels and qf_dates_raw.strip():
            qf_date_error = "No valid dates found. Check format (29-Sep-25, 29/09/2025, 2025-09-29 all work)."

    elif qf_date_mode == "Date range":
        col_from, col_to = st.columns(2)
        with col_from:
            qf_start = st.date_input("From", value=date.today() - timedelta(days=30), key="qf_range_start")
        with col_to:
            qf_end = st.date_input("To", value=last_trading_day(), key="qf_range_end")

        if qf_start and qf_end:
            if qf_end < qf_start:
                qf_date_error = "End date must be after start date."
            else:
                bdays = pd.bdate_range(start=qf_start, end=qf_end)
                qf_dates_with_labels = [(d.strftime('%d-%b-%Y'), d.date()) for d in bdays]
                st.caption(f"Will fetch **{len(qf_dates_with_labels)} trading days** ({qf_start.strftime('%d-%b-%Y')} → {qf_end.strftime('%d-%b-%Y')})")

    fetch_quick = st.button("🔄 Fetch Closing Prices", type="primary",
                             use_container_width=True, key="btn_quick")

    if fetch_quick:
        qf_symbols = parse_quick_symbols(qf_symbols_raw)
        if not qf_symbols:
            st.error("⚠️ Enter at least one symbol above.")
        elif qf_date_error:
            st.error(f"⚠️ {qf_date_error}")
        elif not qf_dates_with_labels:
            st.error("⚠️ Select or enter at least one date.")
        else:
            run_fetch_and_display(qf_symbols, qf_dates_with_labels, fill_holidays, adjusted, auto_date=False)


# ── FOOTER ───────────────────────────────────────────────────
st.markdown("---")
st.caption("Bhavcopy v1.3  |  NSE stocks + NSE/BSE indexes via Yahoo Finance (unadjusted by default)  |  Blanks = market holiday or weekend  |  Built for Motilal Oswal Research")
